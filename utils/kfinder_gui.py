"""
kfinder_gui.py
==============

Поисковая утилита K-Finder для быстрого доступа к заказам и лазерным файлам.

Назначение
----------
Программа решает три независимых задачи:

1. Поиск папки заказа по K-номеру (K-Finder).
   Каждый заказ имеет папку на сервере G:\\Auftragsdokumente\\<год>\\Kxxxxx.
   Индекс папок хранится в k_index.json.

2. Поиск DXF/DWG-файлов лазерной резки по K-номеру или DXF-номеру.
   DXF-номер — это независимый номер листа лазерной резки; один лист может
   содержать детали из нескольких заказов. Данные берутся из Excel-таблицы
   DXF-2017.xlsm. Индекс хранится в dxf_excel_index.json и dxf_index.json.

3. Поиск заказа по Apparate-Nr. (серийному номеру аппарата).
   Связь серийный номер → K-номер хранится в appnr_index.json.

Архитектура (порядок объявления)
---------------------------------
Конфигурация:
  AppConfig       — настройки K-поиска (пути, параметры обхода)
  DXFConfig       — настройки DXF-модуля (пути к Excel и JSON-индексам)
  AppNrConfig     — настройки модуля Apparate-Nr.

Модели данных:
  KEntry          — запись индекса K-папок
  DXFRange        — диапазон DXF-номеров → папка на диске
  DXFExcelRecord  — строка из Excel-таблицы лазерной резки
  DXFFileRecord   — запись об основном DWG-файле на диске
  DXFSearchResult — объединённый результат поиска DXF
  AppNrRecord     — запись серийный номер → K-номер

Репозитории (бизнес-логика):
  KIndex          — in-memory кэш K-папок, хвостовое и полное обновление
  SearchService   — поиск K-записи (индекс → диск → кэш отсутствия)
  DXFRepository   — три индекса (диапазоны, Excel, файлы) + поиск
  AppNrRepository — индекс серийных номеров + поиск

GUI:
  _make_gen_button  — фабрика стилизованной GenButton (AT-CAD-стиль)
  _static_box_sizer — фабрика StaticBoxSizer (AT-CAD-стиль)
  ResultsDialog     — окно результатов K-поиска
  DXFResultsDialog  — окно результатов DXF-поиска
  ServiceDialog     — служебное окно (индексация)
  AboutDialog       — информация о программе
  KFinderFrame      — главное окно

Сборка .exe
-----------
pyinstaller --noconfirm --clean --onefile --windowed --icon=kfinder.ico --name=kfinder --hidden-import=openpyxl kfinder_gui.py
"""

from __future__ import annotations
import json
import logging
import os
import re
import subprocess
import sys
import threading
from dataclasses import dataclass, asdict
from datetime import datetime
from pathlib import Path
from typing import Optional, Callable
import wx
from openpyxl import load_workbook
from wx.lib.buttons import GenButton

# ============================================================
# Рабочая папка приложения
# ============================================================

def _app_dir() -> Path:
    """
    Возвращает папку, в которой лежит исполняемый файл.
    Корректно работает как для .py-скрипта, так и для .exe (PyInstaller).
    """
    if getattr(sys, "frozen", False):
        return Path(sys.executable).resolve().parent
    return Path(__file__).resolve().parent


APP_DIR = _app_dir()

# ============================================================
# Логирование
# ============================================================
#
# Используем именованный logger, чтобы не перебивать basicConfig
# других модулей (если программа встроена в AT-CAD).
# Логируем только ошибки — информационные сообщения идут в строку
# статуса главного окна.

_LOG_FILE = APP_DIR / "kfinder.log"
logger = logging.getLogger("kfinder")
logger.propagate = False  # не передавать выше по иерархии логгеров

if not logger.handlers:
    try:
        _h = logging.FileHandler(_LOG_FILE, encoding="utf-8")
        _h.setFormatter(logging.Formatter("%(asctime)s [%(levelname)s] %(message)s"))
        _h.setLevel(logging.ERROR)
        logger.addHandler(_h)
    except OSError:
        # Если лог-файл недоступен (права, readonly-диск) — просто молчим
        pass

logger.setLevel(logging.ERROR)

# ============================================================
# Строки интерфейса
# ============================================================
# Все тексты в одном месте — легко найти, легко перевести.
# Форматные вставки: {code}, {count}, {path}, {query} и т.п.

TXT = {
    # Главное окно
    "app_title":               "K-Finder  —  Auftragsverzeichnis",
    "app_subtitle":            "Auftragsverzeichnis-Suche",
    "input_box":               "Bitte K-Nr., DXF-Nr. o. App.Nr. auswählen und einfügen",
    "actions_box":             "Suchen & Öffnen",
    "service_box":             "Service",
    "meta_box":                "Indexinformationen",
    "status_box":              "Status",

    # Кнопки главного окна
    "search_show":             "🔍  Ordner anzeigen",
    "open_folder":             "📁  Auftragsordner öffnen",
    "open_sketch":             "✏️  RP - dxf -Skizzen",
    "open_dwg":                "📐  ABW-K…dwg öffnen",
    "open_dxf":                "📄  DXF/DWG anzeigen",
    "close_program":           "✖  Programm beenden",

    # Нижние кнопки
    "service_button":          "Service",
    "about_button":            "Info",

    # Статус
    "status_ready_short":      "Bereit.",
    "status_root_warn":        "⚠ Root-Verzeichnis nicht verfügbar",
    "status_no_hits":          "Keine Treffer im Index.",
    "status_rebuild_running":  "Index wird neu aufgebaut…",
    "status_rebuild_done":     "Index neu aufgebaut. Einträge: {count}",
    "status_rebuild_error":    "Fehler beim Neuaufbau",
    "status_found_one":        "1 Treffer: {code}",
    "status_found_many":       "{count} Treffer gefunden.",
    "status_found":            "{code} gefunden.",
    "status_found_with_serial":"{code} gefunden. App.-Nr.: {serials}",
    "status_not_found":        "{code} nicht gefunden.",
    "status_no_folder":        "⚠ {code}: kein eigenes Verzeichnis",
    "status_searching":        "Suche {code}…",
    "status_start_update":     "Automatische Aktualisierung beim Start…",
    "status_ready":            "Bereit.",

    # Режимы поиска
    "input_hint":              "Suchtyp und Nummer eingeben:",
    "input_hint_k":            "K-Nummer:",
    "input_hint_dxf":          "DXF-Nummer:",
    "input_hint_app":          "Apparate-Nr.:",
    "search_mode_k":           "K",
    "search_mode_dxf":         "DXF",
    "search_mode_app":         "App. Nr.",
    "search_mode_menu_title":  "Suchtyp wählen",

    # Сообщения
    "msg_input_required":      "Bitte Auftragsnummer eingeben.",
    "msg_input_error":         "Eingabefehler",
    "msg_hint":                "Hinweis",
    "msg_warning":             "Warnung",
    "msg_error":               "Fehler",
    "msg_not_found_title":     "Nicht gefunden",
    "msg_no_hits_title":       "Keine Treffer",
    "msg_no_folder_title":     "Kein Verzeichnis",
    "msg_select_entry":        "Bitte einen Eintrag auswählen.",
    "msg_folder_missing":      "Ordner nicht gefunden:\n{path}",
    "msg_file_missing":        "Datei nicht gefunden:\n{path}",
    "msg_root_unavailable":    "Root-Verzeichnis nicht erreichbar:\n{path}",
    "msg_not_found_full":      "{code} wurde weder im Index noch auf dem Datenträger gefunden.",
    "msg_no_hits":             "Keine Treffer für '{query}'.",
    "msg_no_folder_single":    "{code} hat keine eigene Verzeichnisstruktur.",
    "msg_no_folder_full": (
        "{code} ist im Index vorhanden, hat aber kein eigenes Verzeichnis.\n"
        "Möglicherweise wurde dieser Auftrag unter einer anderen Nummer abgelegt."
    ),
    "msg_mode_not_supported":  "Diese Aktion ist für den gewählten Suchtyp nicht verfügbar.",

    # Сообщения — DXF-режим
    "msg_dxf_input_error":      "Zulässige Eingabe: nur Ziffern der DXF-Nummer, z. B. 11601 oder 116.",
    "msg_dxf_not_found_title":  "DXF nicht gefunden",
    "msg_dxf_not_found":        "Für DXF '{query}' wurden keine Daten gefunden.",

    # Сообщения — App.Nr.-режим
    "msg_app_input_error":      "Zulässige Eingabe: Apparate-Nr. oder deren Anfang, z. B. 1234 oder 1234.1-2.",
    "msg_app_not_found_title":  "Apparate-Nr. nicht gefunden",
    "msg_app_not_found":        "Für Apparate-Nr. '{query}' wurde keine Zuordnung gefunden.",
    "msg_app_multi_title":      "Mehrere Apparate-Nr. gefunden",
    "msg_app_multi":            "Für Apparate-Nr. '{query}' wurden mehrere Treffer gefunden.",

    # Диалог результатов K-поиска
    "results_box":             "Gefundene Aufträge",
    "results_title":           "Gefundene Aufträge",
    "actions_title":           "Aktionen",
    "close_dialog":            "✖  Schließen",
    "table_order":             "Auftrag",
    "table_year":              "Jahr",
    "table_folder_exists":     "Ordner vorhanden",
    "table_folder":            "Auftragsordner",
    "table_sketch":            "RP-DXF-Skizzen",
    "table_dwg":               "DWG-Datei",
    "table_has_folder_yes":    "✅",
    "table_has_folder_no":     "⚠ kein Ordner",

    # Диалог результатов DXF-поиска
    "dxf_results_title":       "DXF/DWG-Ergebnisse",
    "dxf_box":                 "DXF/DWG-Daten",
    "dxf_actions_box":         "Aktionen",
    "dxf_no_hits_title":       "Keine DXF-Daten",
    "dxf_no_hits":             "Keine DXF/DWG-Daten für '{code}' gefunden.",
    "dxf_open_folder":         "Ordner öffnen",
    "dxf_open_file":           "DWG öffnen",
    "dxf_save_file":           "In Datei speichern",
    "dxf_close":               "Schließen",
    "dxf_select_entry":        "Bitte eine Zeile auswählen.",
    "dxf_file_missing":        "DWG-Datei nicht gefunden:\n{path}",
    "dxf_folder_missing":      "Ordner nicht gefunden:\n{path}",
    "dxf_save_title":          "Ergebnis speichern",
    "dxf_save_done":           "Datei wurde gespeichert:\n{path}",
    "dxf_col_no":              "DXF",
    "dxf_col_k":               "K-Nr.",
    "dxf_col_wst":             "Werkstoff",
    "dxf_col_dicke":           "Dicke, mm",
    "dxf_col_area":            "A Kn brutto, qm",
    "dxf_col_length":          "Länge Zuschnitt, mm",
    "dxf_col_price":           "Preis/Länge €",
    "dxf_col_file":            "DWG",

    # Диалог результатов App.Nr.-поиска
    "app_results_title":       "Apparate-Nr.-Ergebnisse",
    "app_box":                 "Gefundene Apparate",
    "app_actions_box":         "Aktionen",
    "app_select_entry":        "Bitte eine Zeile auswählen.",
    "app_col_serial":          "App.-Nr.",
    "app_col_prefix":          "Präfix",
    "app_col_k":               "K-Nr.",
    "app_col_folder_exists":   "Ordner vorhanden",
    "app_col_folder":          "Auftragsordner",

    # Служебный диалог
    "service_dialog_title":    "Service",
    "service_info_box":        "Indexinformationen",
    "service_actions_box":     "Aktionen",
    "service_update_partial":  "Teilaktualisierung",
    "service_update_full":     "Vollständige Neuindizierung",
    "service_close":           "Schließen",
    "service_partial_confirm": "Teilaktualisierung des Indexes ausführen?",
    "service_full_confirm":    "Vollständige Neuindizierung ausführen?",
    "service_update_running":  "Teilaktualisierung läuft…",
    "service_rebuild_running": "Vollständige Neuindizierung läuft…",
    "service_partial_done":    "Aktualisiert. Einträge: {count}",
    "service_full_done":       "Neuindizierung abgeschlossen. Einträge: {count}",
    "service_update_error":    "Fehler bei der Teilaktualisierung",
    "service_rebuild_error":   "Fehler bei der Neuindizierung",

    # Обновление индекса (общие)
    "update_confirm_title":    "Daten aktualisieren",
    "update_confirm":          "Indexdaten aktualisieren?",
    "update_already_running":  "Die Aktualisierung läuft bereits.",
    "update_done_title":       "Fertig",
    "update_done_msg":         "Indexdaten wurden aktualisiert.\nEinträge: {count}",
    "msg_rebuild_confirm_title": "Index neu aufbauen",
    "msg_rebuild_confirm":       "Index vollständig neu aufbauen?",
    "msg_rebuild_already_running": "Der Indexaufbau läuft bereits.",
    "msg_rebuild_done_title":    "Fertig",
    "msg_rebuild_done":          "Index neu aufgebaut.\nEinträge: {count}",

    # About
    "about_title":             "Über K-Finder",
    "about_text_title":        "K-Finder",
    "about_text_subtitle":     "Suche und Navigation für Auftragsdaten",
    "about_text_body": (
        "Das Programm ermöglicht den schnellen Zugriff auf\n"
        "Auftragsordner, Skizzenordner und DWG-Dateien.\n\n"
        "Unterstützte Suchtypen:\n"
        "• K-Nummer\n"
        "• DXF-Nummer\n"
        "• Apparate-Nr.\n"
    ),
    "about_text_footer": (
        "Autor: A. Tutubalin\n"
        "Version: 3.2\n"
        "© 2026"
    ),
    "about_ok":                "OK",
}


# ============================================================
# КОНФИГУРАЦИЯ
# ============================================================

@dataclass(frozen=True)
class AppConfig:
    """
    Неизменяемая конфигурация K-модуля.

    root_dir          — корень папок заказов на сервере.
    index_file        — путь к JSON-файлу индекса K-папок.
    start_year        — самый ранний год для полного сканирования.
    sketch_folder_name— имя подпапки со скетчами внутри папки заказа.
    auto_open_single  — открыть папку автоматически если найден 1 результат.
    auto_show_single  — показать ResultsDialog если найден 1 результат.
    live_search_if_missing — искать на диске, если номера нет в индексе.
    window_size       — размер главного окна (ширина, высота).
    service_window_size — размер служебного окна.
    tail_backtrack    — на сколько номеров назад проверять при хвостовом обновлении.
    tail_years_to_scan— сколько последних лет сканировать при хвостовом обновлении.
    auto_update_on_start — запускать хвостовое обновление при старте программы.
    """
    root_dir:               Path  = Path(r"G:\Auftragsdokumente")
    index_file:             Path  = APP_DIR / "k_index.json"
    start_year:             int   = 2011
    sketch_folder_name:     str   = "RP - dxf -Skizzen"
    auto_open_single:       bool  = False
    auto_show_single:       bool  = True
    live_search_if_missing: bool  = True
    window_size:            tuple = (410, 490)
    service_window_size:    tuple = (500, 360)
    tail_backtrack:         int   = 50
    tail_years_to_scan:     int   = 2
    auto_update_on_start:   bool  = True


@dataclass(frozen=True)
class DXFConfig:
    """
    Неизменяемая конфигурация DXF-модуля.

    dxf_root_dir       — корень папок DXF/DWG-файлов на диске.
    excel_file         — путь к Excel-таблице лазерной резки (DXF-2017.xlsm).
    ranges_json        — JSON с картой диапазонов DXF-номеров → папка.
    excel_index_json   — JSON-индекс строк Excel, ключ: K-номер.
    files_index_json   — JSON-индекс основных DWG-файлов, ключ: DXF-номер.
    min_dxf_no         — первый DXF-номер в Excel (7801).
    max_dxf_no         — верхняя граница для полного сканирования.
    file_tail_backtrack— откат назад при хвостовом обновлении файлового индекса.
    file_tail_forward_scan — сколько номеров вперёд проверять при хвостовом обновлении.
    """
    dxf_root_dir:       Path  = Path(r"G:\Drawing\DXF-LASER")
    excel_file:         Path  = Path(r"G:\Drawing\DXF-LASER\DXF-2017.xlsm")
    ranges_json:        Path  = APP_DIR / "dxf_ranges.json"
    excel_index_json:   Path  = APP_DIR / "dxf_excel_index.json"
    files_index_json:   Path  = APP_DIR / "dxf_index.json"
    min_dxf_no:         int   = 7801
    max_dxf_no:         int   = 20000
    file_tail_backtrack:     int = 100
    file_tail_forward_scan:  int = 700


@dataclass(frozen=True)
class AppNrConfig:
    """
    Конфигурация модуля Apparate-Nr. (серийные номера аппаратов).

    excel_file  — Excel-файл с колонками: A = серийный номер, E = K-номер.
    index_json  — JSON-индекс серийных номеров.
    """
    excel_file: Path = Path(r"G:\Auftragsdokumente\Apparate-Nr.xlsx")
    index_json: Path = APP_DIR / "appnr_index.json"


# Глобальные экземпляры конфигураций
APP_CONFIG   = AppConfig()
DXF_CONFIG   = DXFConfig()
APPNR_CONFIG = AppNrConfig()


# ============================================================
# МОДЕЛИ ДАННЫХ
# ============================================================

@dataclass
class KEntry:
    """
    Запись индекса папок заказов.

    k_code      — код заказа в формате K12345.
    year        — год (из структуры каталога). 0 = неизвестен.
    folder_path — полный путь к папке заказа.
    sketch_path — полный путь к подпапке с скетчами (RP - dxf -Skizzen).
    dwg_path    — полный путь к основному DWG-файлу в папке скетчей.
    has_folder  — True если папка заказа реально существует на диске.
                  False для записей-заглушек (пропущенные номера).
    """
    k_code:      str
    year:        int
    folder_path: str
    sketch_path: str
    dwg_path:    str
    has_folder:  bool = True

    @property
    def folder_exists(self) -> bool:
        """Проверяет существование папки заказа в реальном времени."""
        return Path(self.folder_path).is_dir()

    @property
    def sketch_exists(self) -> bool:
        """Проверяет существование папки скетчей в реальном времени."""
        return Path(self.sketch_path).is_dir()

    @property
    def dwg_exists(self) -> bool:
        """Проверяет существование основного DWG-файла в реальном времени."""
        return Path(self.dwg_path).is_file()

    def to_dict(self) -> dict:
        return asdict(self)

    @staticmethod
    def from_dict(d: dict) -> "KEntry":
        return KEntry(
            k_code=d["k_code"],
            year=int(d["year"]),
            folder_path=d["folder_path"],
            sketch_path=d["sketch_path"],
            dwg_path=d["dwg_path"],
            has_folder=bool(d.get("has_folder", True)),
        )


@dataclass
class DXFRange:
    """
    Диапазон DXF-номеров, соответствующий одной папке на диске.

    min_no      — первый номер диапазона (включительно).
    max_no      — последний номер диапазона (включительно).
    folder_name — имя подпапки относительно dxf_root_dir.

    Пример: DXFRange(7801, 9500, "2019") означает, что файлы
    7801.dwg … 9500.dwg лежат в G:\\Drawing\\DXF-LASER\\2019\\.
    """
    min_no:      int
    max_no:      int
    folder_name: str

    def to_dict(self) -> dict:
        return asdict(self)

    @staticmethod
    def from_dict(d: dict) -> "DXFRange":
        return DXFRange(
            min_no=int(d["min_no"]),
            max_no=int(d["max_no"]),
            folder_name=str(d["folder_name"]),
        )


@dataclass
class DXFExcelRecord:
    """
    Одна строка из Excel-таблицы DXF-2017.xlsm (лист Tabelle1).

    dxf_no              — номер DXF-листа (колонка A).
    k_num               — K-номер заказа (колонка B, нормализован к формату K12345).
    schluessel           — шифр/ключ детали (колонка C).
    wst                 — материал (колонка D).
    dicke_mm            — толщина листа в мм (колонка E).
    a_kn_brutto_qm      — площадь брутто в кв.м (колонка I).
    laenge_zuschnitt_mm — длина заготовки в мм (колонка K).
    preis_pro_laenge_eur— цена за длину в евро (колонка N).
    bemerkung           — примечания (колонка S).
    """
    dxf_no:               int
    k_num:                str
    schluessel:           str
    wst:                  str
    dicke_mm:             Optional[float]
    a_kn_brutto_qm:       Optional[float]
    laenge_zuschnitt_mm:  Optional[float]
    preis_pro_laenge_eur: Optional[float]
    bemerkung:            str

    def to_dict(self) -> dict:
        return asdict(self)

    @staticmethod
    def from_dict(d: dict) -> "DXFExcelRecord":
        return DXFExcelRecord(
            dxf_no=int(d["dxf_no"]),
            k_num=str(d["k_num"]),
            schluessel=str(d.get("schluessel", "")),
            wst=str(d.get("wst", "")),
            dicke_mm=d.get("dicke_mm"),
            a_kn_brutto_qm=d.get("a_kn_brutto_qm"),
            laenge_zuschnitt_mm=d.get("laenge_zuschnitt_mm"),
            preis_pro_laenge_eur=d.get("preis_pro_laenge_eur"),
            bemerkung=str(d.get("bemerkung", "")),
        )


@dataclass
class DXFFileRecord:
    """
    Запись о DWG-файле на диске.

    dxf_no        — номер DXF-листа.
    folder_name   — имя подпапки (из карты диапазонов).
    folder_path   — полный путь к папке с файлом.
    main_dwg_path — полный путь к основному DWG-файлу (xxxxx.dwg без суффиксов).
    has_main_dwg  — True если файл реально существует на диске.
    """
    dxf_no:       int
    folder_name:  str
    folder_path:  str
    main_dwg_path: str
    has_main_dwg: bool

    def to_dict(self) -> dict:
        return asdict(self)

    @staticmethod
    def from_dict(d: dict) -> "DXFFileRecord":
        return DXFFileRecord(
            dxf_no=int(d["dxf_no"]),
            folder_name=str(d["folder_name"]),
            folder_path=str(d["folder_path"]),
            main_dwg_path=str(d["main_dwg_path"]),
            has_main_dwg=bool(d["has_main_dwg"]),
        )


@dataclass
class DXFSearchResult:
    """
    Объединённый результат поиска: данные из Excel + путь к файлу.
    Используется для отображения в DXFResultsDialog.
    """
    dxf_no:               int
    k_num:                str
    wst:                  str
    dicke_mm:             Optional[float]
    a_kn_brutto_qm:       Optional[float]
    laenge_zuschnitt_mm:  Optional[float]
    preis_pro_laenge_eur: Optional[float]
    main_dwg_path:        str
    has_main_dwg:         bool

    def to_dict(self) -> dict:
        return asdict(self)


@dataclass
class AppNrRecord:
    """
    Запись индекса серийных номеров аппаратов.

    serial_no     — полный серийный номер (например, "1234.1-2").
    serial_prefix — числовой префикс до точки (например, "1234").
    k_code        — связанный K-номер заказа.
    """
    serial_no:     str
    serial_prefix: str
    k_code:        str

    def to_dict(self) -> dict:
        return asdict(self)

    @staticmethod
    def from_dict(d: dict) -> "AppNrRecord":
        return AppNrRecord(
            serial_no=str(d["serial_no"]),
            serial_prefix=str(d["serial_prefix"]),
            k_code=str(d["k_code"]),
        )


# ============================================================
# РЕПОЗИТОРИИ
# ============================================================

class KIndex:
    """
    In-memory кэш индекса папок заказов K-Finder.

    Данные хранятся в k_index.json и загружаются при старте в _items.
    Все операции чтения (find_exact, find_partial) работают из памяти.
    Запись на диск происходит только при save_entries / update_entry.

    Два режима обновления:
    - update_tail() — быстрое хвостовое обновление (последние N лет).
    - rebuild()     — полное перестроение по всем годам.

    Формат JSON:
    {
      "_meta": { "root": "...", "start_year": 2011, "generated_at": "...", "count": N },
      "items": { "K12345": { "k_code": "K12345", "year": 2023, ... }, ... }
    }
    """

    # K-код: ровно буква K и 5 цифр
    FULL_RE    = re.compile(r"^K\d{5}$", re.IGNORECASE)
    # Допустимый частичный запрос: опциональная K и от 1 до 5 цифр
    PARTIAL_RE = re.compile(r"^[Kk]?\d{1,5}$")

    def __init__(self, cfg: AppConfig):
        self.cfg = cfg
        self._items: dict[str, KEntry] = {}
        self._meta:  dict              = {}
        self._load_cache()

    # ------------------------------------------------------------------
    # Кэш и JSON
    # ------------------------------------------------------------------

    def _load_cache(self) -> None:
        """Загружает индекс с диска в память."""
        data = self._load_raw()
        self._meta  = data.get("_meta", {})
        self._items = {}
        for code, d in data.get("items", {}).items():
            try:
                self._items[code] = KEntry.from_dict(d)
            except (KeyError, ValueError, TypeError):
                logger.error(f"KIndex._load_cache: некорректная запись {code}")

    def reload(self) -> None:
        """Принудительно перечитывает индекс с диска."""
        self._load_cache()

    def _load_raw(self) -> dict:
        if not self.cfg.index_file.is_file():
            return {"_meta": {}, "items": {}}
        try:
            with self.cfg.index_file.open("r", encoding="utf-8") as f:
                data = json.load(f)
            if not isinstance(data, dict) or "items" not in data:
                return {"_meta": {}, "items": {}}
            return data
        except (OSError, json.JSONDecodeError) as e:
            logger.error(f"KIndex._load_raw: {e}")
            return {"_meta": {}, "items": {}}

    def _save_raw(self, data: dict) -> None:
        self.cfg.index_file.parent.mkdir(parents=True, exist_ok=True)
        with self.cfg.index_file.open("w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)

    def load_entries(self) -> dict[str, KEntry]:
        """Возвращает копию текущего кэша (для безопасного внешнего доступа)."""
        return dict(self._items)

    def save_entries(self, items: dict[str, KEntry]) -> None:
        """Сохраняет переданный словарь в кэш и на диск, обновляет метаданные."""
        self._items = dict(items)
        self._meta  = {
            "root":         str(self.cfg.root_dir),
            "start_year":   self.cfg.start_year,
            "generated_at": datetime.now().isoformat(timespec="seconds"),
            "count":        len(self._items),
        }
        payload = {
            "_meta": self._meta,
            "items": {k: v.to_dict() for k, v in sorted(self._items.items())},
        }
        self._save_raw(payload)

    def get_meta(self) -> dict:
        """Возвращает копию метаданных индекса."""
        return dict(self._meta)

    def update_entry(self, entry: KEntry) -> None:
        """Добавляет или заменяет одну запись и сохраняет индекс на диск."""
        self._items[entry.k_code] = entry
        self.save_entries(self._items)

    # ------------------------------------------------------------------
    # Вспомогательные методы
    # ------------------------------------------------------------------

    def _current_year(self) -> int:
        return datetime.now().year

    def is_root_available(self) -> bool:
        """Проверяет доступность корневой папки заказов на сервере."""
        return self.cfg.root_dir.exists() and self.cfg.root_dir.is_dir()

    def normalize_full(self, text: str) -> str:
        """
        Приводит ввод к формату K12345.
        Принимает: '12345' → 'K12345', 'k12345' → 'K12345'.
        Бросает ValueError при некорректном формате.
        """
        s = text.strip().upper()
        if re.fullmatch(r"\d{5}", s):
            return f"K{s}"
        if self.FULL_RE.fullmatch(s):
            return s
        raise ValueError("Eingabe: Kxxxxx oder 5 Ziffern")

    def normalize_partial(self, text: str) -> str:
        """
        Приводит частичный запрос к верхнему регистру.
        Бросает ValueError при недопустимом формате.
        """
        s = text.strip().upper()
        if not s:
            return ""
        if not self.PARTIAL_RE.fullmatch(s):
            raise ValueError("Zulässige Eingabe: K20500, 20500, K205, 205 usw.")
        return s

    def is_full_code(self, text: str) -> bool:
        """True если текст является полным K-кодом (K12345 или 5 цифр)."""
        s = text.strip().upper()
        return bool(self.FULL_RE.fullmatch(s) or re.fullmatch(r"\d{5}", s))

    def _code_to_num(self, k_code: str) -> int:
        """K12345 → 12345"""
        return int(k_code[1:])

    def _num_to_code(self, num: int) -> str:
        """12345 → 'K12345'"""
        return f"K{num:05d}"

    def _max_known_number(self) -> int:
        """Возвращает наибольший числовой K-номер в кэше."""
        nums = [self._code_to_num(code) for code in self._items if self.FULL_RE.fullmatch(code)]
        return max(nums) if nums else 0

    def _make_entry(self, k_code: str, year: int, folder: Path) -> KEntry:
        """Создаёт KEntry для найденной папки заказа."""
        sketch = folder / self.cfg.sketch_folder_name
        dwg    = sketch / f"ABW-{k_code}.dwg"
        return KEntry(
            k_code=k_code,
            year=year,
            folder_path=str(folder),
            sketch_path=str(sketch),
            dwg_path=str(dwg),
            has_folder=True,
        )

    def _make_missing_entry(self, k_code: str) -> KEntry:
        """Создаёт запись-заглушку для номера, папка которого не найдена."""
        return KEntry(
            k_code=k_code,
            year=0,
            folder_path="",
            sketch_path="",
            dwg_path="",
            has_folder=False,
        )

    # ------------------------------------------------------------------
    # Сканирование диска
    # ------------------------------------------------------------------

    def _scan_years_for_kfolders(
        self,
        start_year: int,
        end_year: int,
        progress_cb: Optional[Callable[[str], None]] = None,
    ) -> dict[str, KEntry]:
        """
        Один массовый проход по годовым папкам на сервере.

        Алгоритм:
        - Для каждого года обходит дерево папок через os.walk.
        - Если папка называется Kxxxxx — добавляет в результат и
          запрещает дальнейший спуск внутрь (dirnames[:] = []).
        - Вызывает progress_cb("2023: 47 Ordner") после каждого года.

        Возвращает словарь { "K12345": KEntry, ... }.
        """
        found: dict[str, KEntry] = {}

        for year in range(start_year, end_year + 1):
            year_path  = self.cfg.root_dir / str(year)
            count_year = 0

            if year_path.is_dir():
                for dirpath, dirnames, _ in os.walk(year_path):
                    current = Path(dirpath)
                    name    = current.name.upper()

                    if self.FULL_RE.fullmatch(name):
                        found[name] = self._make_entry(name, year, current)
                        dirnames[:] = []  # не заходим внутрь K-папки
                        count_year += 1

            if progress_cb:
                progress_cb(f"{year}: {count_year} Ordner")

        return found

    # ------------------------------------------------------------------
    # Поиск
    # ------------------------------------------------------------------

    def find_exact(self, k_code: str) -> Optional[KEntry]:
        """Точный поиск по коду в кэше. O(1)."""
        return self._items.get(k_code)

    def find_partial(self, query: str) -> list[KEntry]:
        """
        Частичный поиск: возвращает все записи, код которых начинается
        с указанного префикса. Сортирует по году и коду.
        """
        q = self.normalize_partial(query)
        if not q:
            return []

        q_digits = q[1:] if q.startswith("K") else q
        results = [
            e for code, e in self._items.items()
            if code.startswith(q) or code[1:].startswith(q_digits)
        ]
        results.sort(key=lambda e: (e.year, e.k_code))
        return results

    def search_on_disk(self, k_code: str) -> Optional[KEntry]:
        """
        Точечный поиск одного K-кода по всем годовым папкам.
        Используется только если запись отсутствует в кэше.
        Медленнее чем find_exact, но находит новые папки.
        """
        if not self.is_root_available():
            raise FileNotFoundError(f"Root nicht verfügbar: {self.cfg.root_dir}")

        for year in range(self.cfg.start_year, self._current_year() + 1):
            year_path = self.cfg.root_dir / str(year)
            if not year_path.is_dir():
                continue

            for dirpath, dirnames, _ in os.walk(year_path):
                current = Path(dirpath)
                name    = current.name.upper()

                if name == k_code and current.is_dir():
                    return self._make_entry(k_code, year, current)

                # Не заходим внутрь K-папок (оптимизация)
                if self.FULL_RE.fullmatch(name):
                    dirnames[:] = []

        return None

    # ------------------------------------------------------------------
    # Обновление индекса
    # ------------------------------------------------------------------

    def update_tail(
        self,
        backtrack: int = 100,
        tail_years_to_scan: int = 2,
        progress_cb: Optional[Callable[[str], None]] = None,
    ) -> int:
        """
        Инкрементальное (хвостовое) обновление индекса.

        Алгоритм:
        1. Берём наибольший известный K-номер из кэша.
        2. Сканируем только последние tail_years_to_scan лет —
           один проход вместо тысяч точечных поисков.
        3. Обновляем хвостовой диапазон [max_known - backtrack .. max_found]:
           - Если папка найдена — обновляем запись.
           - Если папка не найдена:
             * Для записей с has_folder=True (найденных в прошлом) — НЕ трогаем,
               они могут быть в более ранних годах за пределами хвоста.
             * Для записей с has_folder=False или отсутствующих — ставим заглушку.
        4. Обрезаем хвост: убираем заглушки выше последнего реального номера.
        5. Сохраняем результат.

        Ключевое отличие от rebuild(): сохраняет старые записи has_folder=True
        вне диапазона хвоста неизменными.
        """
        if not self.is_root_available():
            raise FileNotFoundError(f"Root nicht verfügbar: {self.cfg.root_dir}")

        items      = dict(self._items)
        max_known  = self._max_known_number()

        # Если индекс пуст — делаем полное перестроение
        if max_known <= 0:
            return self.rebuild(progress_cb=progress_cb)

        current_year     = self._current_year()
        start_year_scan  = max(self.cfg.start_year, current_year - tail_years_to_scan + 1)

        # Один проход по последним годам
        found_recent = self._scan_years_for_kfolders(
            start_year=start_year_scan,
            end_year=current_year,
            progress_cb=progress_cb,
        )

        found_recent_nums = [
            self._code_to_num(code)
            for code in found_recent
            if self.FULL_RE.fullmatch(code)
        ]

        max_found_recent = max(found_recent_nums) if found_recent_nums else max_known
        start_num        = max(1, max_known - backtrack)
        end_num          = max(max_known, max_found_recent)

        # Обновляем хвостовой диапазон
        for num in range(start_num, end_num + 1):
            code     = self._num_to_code(num)
            existing = items.get(code)

            if code in found_recent:
                # Папка найдена в последних годах — всегда обновляем
                items[code] = found_recent[code]
            elif existing is not None and existing.has_folder:
                # Папка была найдена раньше (в более ранних годах вне хвоста) — не трогаем
                pass
            else:
                # Запись отсутствует или уже была заглушкой — ставим заглушку
                items[code] = self._make_missing_entry(code)

        # Обрезаем хвост выше последнего реального номера
        last_real_num = max(
            (self._code_to_num(code) for code, e in items.items()
             if e.has_folder and self.FULL_RE.fullmatch(code)),
            default=0
        )

        trimmed: dict[str, KEntry] = {
            code: entry
            for code, entry in items.items()
            if not self.FULL_RE.fullmatch(code)
            or self._code_to_num(code) <= last_real_num
        }

        self.save_entries(trimmed)
        return len(trimmed)

    def rebuild(
        self,
        progress_cb: Optional[Callable[[str], None]] = None,
    ) -> int:
        """
        Полное перестроение индекса по всем годам.

        Сканирует все годовые папки от start_year до текущего года.
        Заполняет пропуски в диапазоне номеров заглушками has_folder=False —
        чтобы повторный поиск по «несуществующим» номерам не шёл на диск.
        """
        if not self.is_root_available():
            raise FileNotFoundError(f"Root nicht verfügbar: {self.cfg.root_dir}")

        found_items = self._scan_years_for_kfolders(
            start_year=self.cfg.start_year,
            end_year=self._current_year(),
            progress_cb=progress_cb,
        )

        found_nums = [
            self._code_to_num(code)
            for code in found_items
            if self.FULL_RE.fullmatch(code)
        ]

        if not found_nums:
            self.save_entries({})
            return 0

        min_num = min(found_nums)
        max_num = max(found_nums)

        full_items: dict[str, KEntry] = {}
        for num in range(min_num, max_num + 1):
            code = self._num_to_code(num)
            if code in found_items:
                full_items[code] = found_items[code]
            else:
                full_items[code] = self._make_missing_entry(code)

        self.save_entries(full_items)
        return len(full_items)


class SearchService:
    """
    Сервис поиска K-записи: индекс → диск → кэш отсутствия.

    Логика get_or_search():
    1. Проверяем кэш. Если запись есть (в т.ч. has_folder=False) — возвращаем её.
       НЕ идём на диск для has_folder=False (номер уже проверялся ранее).
    2. Если записи нет и live_search_if_missing=True — ищем на диске.
    3. Если нашли на диске — добавляем в индекс и возвращаем.
    4. Если не нашли — сохраняем заглушку has_folder=False, возвращаем её.
    """

    def __init__(self, cfg: AppConfig, index: KIndex):
        self.cfg   = cfg
        self.index = index

    def get_or_search(self, k_code: str) -> Optional[KEntry]:
        entry = self.index.find_exact(k_code)

        # Запись уже есть в индексе (включая has_folder=False) — не идём на диск
        if entry is not None:
            return entry

        if not self.cfg.live_search_if_missing:
            return None

        # Ищем на диске
        entry = self.index.search_on_disk(k_code)
        if entry:
            self.index.update_entry(entry)
            return entry

        # Не найден — сохраняем заглушку, чтобы следующий запрос не ходил на диск
        missing = KEntry(
            k_code=k_code,
            year=0,
            folder_path="",
            sketch_path="",
            dwg_path="",
            has_folder=False,
        )
        self.index.update_entry(missing)
        return missing


class DXFRepository:
    """
    Главный сервис DXF-модуля. Управляет тремя независимыми индексами:

    1. _ranges       — карта диапазонов DXF-номеров → папка на диске.
                       Источник: лист "Key" в DXF-2017.xlsm.
                       Файл: dxf_ranges.json.

    2. _excel_index  — строки Excel, индексированные по K-номеру.
                       Источник: лист "Tabelle1" в DXF-2017.xlsm.
                       Файл: dxf_excel_index.json.
                       Структура: { "K12345": [ DXFExcelRecord, ... ], ... }

    3. _files_index  — пути к основным DWG-файлам, индексированные по DXF-номеру.
                       Источник: файловая система (только xxxxx.dwg без суффиксов).
                       Файл: dxf_index.json.
                       Структура: { 7801: DXFFileRecord, 7802: ..., ... }

    Дополнительно:
    4. _by_dxf_no   — обратный индекс: DXF-номер → список DXFExcelRecord.
                       Строится в памяти при загрузке excel_index.
                       Нужен для быстрого поиска по DXF-номеру (O(1) вместо O(N)).

    Основной DWG-файл — только xxxxx.dwg (без суффиксов -a, -1 и т.п.).
    Файлы с суффиксами содержат очищенные контуры для CAM и нас не интересуют.
    """

    # Регулярное выражение для основного DWG-файла: только цифры + .dwg
    PRIMARY_DWG_RE = re.compile(r"^(\d+)\.dwg$", re.IGNORECASE)
    FULL_K_RE      = re.compile(r"^K\d{5}$", re.IGNORECASE)

    def __init__(self, cfg: DXFConfig = DXF_CONFIG):
        self.cfg = cfg

        # Три основных индекса
        self._ranges:       list[DXFRange]                     = []
        self._excel_index:  dict[str, list[DXFExcelRecord]]    = {}
        self._files_index:  dict[int, DXFFileRecord]           = {}

        # Обратный индекс для быстрого поиска по DXF-номеру
        self._by_dxf_no:    dict[int, list[DXFExcelRecord]]    = {}

        self.reload_all()

    # ------------------------------------------------------------------
    # Загрузка / перезагрузка
    # ------------------------------------------------------------------

    def reload_all(self) -> None:
        """Загружает все три индекса с диска в память."""
        self._ranges      = self._load_ranges_json()
        self._excel_index = self._load_excel_index_json()
        self._files_index = self._load_files_index_json()
        self._rebuild_dxf_no_index()

    def _rebuild_dxf_no_index(self) -> None:
        """
        Строит обратный индекс _by_dxf_no из _excel_index.
        Вызывается автоматически после каждой загрузки Excel-индекса.
        O(N) по количеству строк Excel.
        """
        self._by_dxf_no = {}
        for records in self._excel_index.values():
            for rec in records:
                self._by_dxf_no.setdefault(rec.dxf_no, []).append(rec)

    # ------------------------------------------------------------------
    # Общие утилиты
    # ------------------------------------------------------------------

    def _normalize_k_num(self, value) -> str:
        """
        Приводит значение ячейки Excel к формату K12345.
        Пустые строки и некорректные значения возвращают ''.
        """
        if value is None:
            return ""
        s = str(value).strip().upper()
        if not s:
            return ""
        if re.fullmatch(r"\d+", s):
            return f"K{int(s):05d}"
        if self.FULL_K_RE.fullmatch(s):
            return s
        return ""

    def _safe_float(self, value) -> Optional[float]:
        """Безопасное преобразование ячейки Excel в float. None при ошибке."""
        if value is None:
            return None
        try:
            return float(value)
        except (TypeError, ValueError):
            return None

    def _ensure_parent(self, path: Path) -> None:
        """Создаёт родительские директории если их нет."""
        path.parent.mkdir(parents=True, exist_ok=True)

    def is_dxf_root_available(self) -> bool:
        """Проверяет доступность корневой папки DXF-файлов."""
        return self.cfg.dxf_root_dir.exists() and self.cfg.dxf_root_dir.is_dir()

    def is_excel_available(self) -> bool:
        """Проверяет доступность Excel-файла."""
        return self.cfg.excel_file.exists() and self.cfg.excel_file.is_file()

    def _get_range_for_dxf(self, dxf_no: int) -> Optional[DXFRange]:
        """Возвращает диапазон для указанного DXF-номера или None."""
        for r in self._ranges:
            if r.min_no <= dxf_no <= r.max_no:
                return r
        return None

    def is_primary_dwg_name(self, filename: str) -> bool:
        """True если имя файла является основным DWG (xxxxx.dwg без суффиксов)."""
        return bool(self.PRIMARY_DWG_RE.fullmatch(filename.strip()))

    # ------------------------------------------------------------------
    # JSON: диапазоны
    # ------------------------------------------------------------------

    def _load_ranges_json(self) -> list[DXFRange]:
        if not self.cfg.ranges_json.is_file():
            return []
        try:
            with self.cfg.ranges_json.open("r", encoding="utf-8") as f:
                data = json.load(f)
            return [DXFRange.from_dict(item) for item in data]
        except (OSError, json.JSONDecodeError, KeyError, TypeError, ValueError) as e:
            logger.error(f"DXFRepository._load_ranges_json: {e}")
            return []

    def _save_ranges_json(self, ranges: list[DXFRange]) -> None:
        self._ensure_parent(self.cfg.ranges_json)
        with self.cfg.ranges_json.open("w", encoding="utf-8") as f:
            json.dump([r.to_dict() for r in ranges], f, ensure_ascii=False, indent=2)

    # ------------------------------------------------------------------
    # JSON: Excel-индекс
    # ------------------------------------------------------------------

    def _load_excel_index_json(self) -> dict[str, list[DXFExcelRecord]]:
        if not self.cfg.excel_index_json.is_file():
            return {}
        try:
            with self.cfg.excel_index_json.open("r", encoding="utf-8") as f:
                data = json.load(f)
            return {
                k_num: [DXFExcelRecord.from_dict(x) for x in items]
                for k_num, items in data.items()
            }
        except (OSError, json.JSONDecodeError, KeyError, TypeError, ValueError) as e:
            logger.error(f"DXFRepository._load_excel_index_json: {e}")
            return {}

    def _save_excel_index_json(self, index_data: dict[str, list[DXFExcelRecord]]) -> None:
        self._ensure_parent(self.cfg.excel_index_json)
        payload = {
            k_num: [item.to_dict() for item in items]
            for k_num, items in sorted(index_data.items())
        }
        with self.cfg.excel_index_json.open("w", encoding="utf-8") as f:
            json.dump(payload, f, ensure_ascii=False, indent=2)

    # ------------------------------------------------------------------
    # JSON: файловый индекс
    # ------------------------------------------------------------------

    def _load_files_index_json(self) -> dict[int, DXFFileRecord]:
        if not self.cfg.files_index_json.is_file():
            return {}
        try:
            with self.cfg.files_index_json.open("r", encoding="utf-8") as f:
                data = json.load(f)
            return {int(key): DXFFileRecord.from_dict(item) for key, item in data.items()}
        except (OSError, json.JSONDecodeError, KeyError, TypeError, ValueError) as e:
            logger.error(f"DXFRepository._load_files_index_json: {e}")
            return {}

    def _save_files_index_json(self, index_data: dict[int, DXFFileRecord]) -> None:
        self._ensure_parent(self.cfg.files_index_json)
        payload = {
            str(dxf_no): item.to_dict()
            for dxf_no, item in sorted(index_data.items())
        }
        with self.cfg.files_index_json.open("w", encoding="utf-8") as f:
            json.dump(payload, f, ensure_ascii=False, indent=2)

    # ------------------------------------------------------------------
    # Построение из Excel — диапазоны (лист Key)
    # ------------------------------------------------------------------

    def rebuild_ranges_from_workbook(self) -> int:
        """
        Строит dxf_ranges.json из листа "Key" в Excel-файле.

        Формат листа Key:
          Строка 1: заголовок (пропускается).
          Столбец A: min_no (начало диапазона).
          Столбец B: max_no (конец диапазона).
          Столбец C: имя подпапки (например "2019").

        Возвращает количество считанных диапазонов.
        """
        if not self.is_excel_available():
            raise FileNotFoundError(f"Excel-Datei nicht gefunden: {self.cfg.excel_file}")

        wb = load_workbook(self.cfg.excel_file, data_only=True, read_only=True, keep_vba=True)
        if "Key" not in wb.sheetnames:
            raise KeyError("Blatt 'Key' wurde in der Excel-Datei nicht gefunden.")

        ws     = wb["Key"]
        result: list[DXFRange] = []
        first  = True

        for row in ws.iter_rows(values_only=True):
            if first:
                first = False
                continue

            min_no, max_no, path_name = (row + (None, None, None))[:3]
            if min_no is None or max_no is None or path_name is None:
                continue

            try:
                result.append(DXFRange(
                    min_no=int(min_no),
                    max_no=int(max_no),
                    folder_name=str(path_name).strip(),
                ))
            except (TypeError, ValueError):
                continue

        wb.close()
        self._save_ranges_json(result)
        self._ranges = result
        return len(result)

    # ------------------------------------------------------------------
    # Построение из Excel — строки (лист Tabelle1)
    # ------------------------------------------------------------------

    def rebuild_excel_index(self) -> int:
        """
        Строит dxf_excel_index.json из листа "Tabelle1" в Excel-файле.

        Колонки Excel (0-based индексы):
          A=0  DXF-номер
          B=1  K-номер (Komm.)
          C=2  Schlüssel
          D=3  Werkstoff
          E=4  Dicke, mm
          I=8  A Kn brutto, qm
          K=10 Länge Zuschnitt, mm
          N=13 Preis/Länge €
          S=18 Bemerkungen

        Строки с пустым DXF- или K-номером пропускаются.
        Внутри каждого K-номера строки сортируются по DXF-номеру.

        После построения обновляет обратный индекс _by_dxf_no.
        Возвращает количество считанных строк.
        """
        if not self.is_excel_available():
            raise FileNotFoundError(f"Excel-Datei nicht gefunden: {self.cfg.excel_file}")

        wb = load_workbook(self.cfg.excel_file, data_only=True, read_only=True, keep_vba=True)
        if "Tabelle1" not in wb.sheetnames:
            raise KeyError("Blatt 'Tabelle1' wurde in der Excel-Datei nicht gefunden.")

        ws           = wb["Tabelle1"]
        index_data:  dict[str, list[DXFExcelRecord]] = {}
        first        = True
        total        = 0

        for row in ws.iter_rows(values_only=True):
            if first:
                first = False
                continue

            # Читаем нужные колонки с защитой от коротких строк
            def _col(i: int):
                return row[i] if len(row) > i else None

            dxf_no_raw = _col(0)
            komm       = _col(1)

            # Пропускаем строки без обоих ключевых полей
            if dxf_no_raw is None or komm is None:
                continue

            try:
                dxf_no_int = int(dxf_no_raw)
            except (TypeError, ValueError):
                continue

            k_num = self._normalize_k_num(komm)
            if not k_num:
                continue

            record = DXFExcelRecord(
                dxf_no=dxf_no_int,
                k_num=k_num,
                schluessel=str(_col(2) or ""),
                wst=str(_col(3) or ""),
                dicke_mm=self._safe_float(_col(4)),
                a_kn_brutto_qm=self._safe_float(_col(8)),
                laenge_zuschnitt_mm=self._safe_float(_col(10)),
                preis_pro_laenge_eur=self._safe_float(_col(13)),
                bemerkung=str(_col(18) or ""),
            )

            index_data.setdefault(k_num, []).append(record)
            total += 1

        wb.close()

        # Сортируем строки внутри каждого K-номера
        for k_num in index_data:
            index_data[k_num].sort(key=lambda x: (x.dxf_no, x.schluessel))

        self._save_excel_index_json(index_data)
        self._excel_index = index_data
        self._rebuild_dxf_no_index()  # обновляем обратный индекс
        return total

    # ------------------------------------------------------------------
    # Построение файлового индекса
    # ------------------------------------------------------------------

    def _build_file_record(self, dxf_no: int) -> DXFFileRecord:
        """
        Строит DXFFileRecord для одного DXF-номера.
        Проверяет наличие файла на диске в реальном времени.
        """
        r = self._get_range_for_dxf(dxf_no)
        if r is None:
            # Номер вне всех известных диапазонов
            return DXFFileRecord(
                dxf_no=dxf_no,
                folder_name="",
                folder_path="",
                main_dwg_path="",
                has_main_dwg=False,
            )

        folder_path = self.cfg.dxf_root_dir / r.folder_name
        main_dwg    = folder_path / f"{dxf_no}.dwg"

        return DXFFileRecord(
            dxf_no=dxf_no,
            folder_name=r.folder_name,
            folder_path=str(folder_path),
            main_dwg_path=str(main_dwg),
            has_main_dwg=main_dwg.is_file(),
        )

    def rebuild_files_index_full(
        self,
        progress_cb: Optional[Callable[[str], None]] = None,
    ) -> int:
        """
        Полная перестройка файлового индекса.
        Проверяет все номера от min_dxf_no до max_dxf_no.
        Требует предварительно загруженных диапазонов (_ranges).
        """
        if not self.is_dxf_root_available():
            raise FileNotFoundError(f"DXF-Root nicht verfügbar: {self.cfg.dxf_root_dir}")

        if not self._ranges:
            self.rebuild_ranges_from_workbook()

        result: dict[int, DXFFileRecord] = {}
        total  = self.cfg.max_dxf_no - self.cfg.min_dxf_no + 1
        done   = 0

        for dxf_no in range(self.cfg.min_dxf_no, self.cfg.max_dxf_no + 1):
            result[dxf_no] = self._build_file_record(dxf_no)
            done += 1
            if progress_cb and (done % 100 == 0 or done == total):
                progress_cb(f"DXF {dxf_no} geprüft ({done}/{total})")

        self._save_files_index_json(result)
        self._files_index = result
        return len(result)

    def update_files_index_tail(
        self,
        backtrack: Optional[int] = None,
        forward_scan: Optional[int] = None,
        progress_cb: Optional[Callable[[str], None]] = None,
    ) -> int:
        """
        Хвостовое обновление файлового индекса.

        Перепроверяет диапазон [max_known - backtrack .. max_known + forward_scan].
        После проверки обрезает хвост: убирает записи выше последнего реального файла.
        """
        if not self.is_dxf_root_available():
            raise FileNotFoundError(f"DXF-Root nicht verfügbar: {self.cfg.dxf_root_dir}")

        if not self._ranges:
            self.rebuild_ranges_from_workbook()

        if not self._files_index:
            return self.rebuild_files_index_full(progress_cb=progress_cb)

        if backtrack is None:
            backtrack = self.cfg.file_tail_backtrack
        if forward_scan is None:
            forward_scan = self.cfg.file_tail_forward_scan

        known_numbers = list(self._files_index.keys())
        max_known     = max(known_numbers) if known_numbers else self.cfg.min_dxf_no
        start_no      = max(self.cfg.min_dxf_no, max_known - backtrack)
        end_no        = min(self.cfg.max_dxf_no,  max_known + forward_scan)

        result = dict(self._files_index)
        total  = end_no - start_no + 1
        done   = 0

        for dxf_no in range(start_no, end_no + 1):
            result[dxf_no] = self._build_file_record(dxf_no)
            done += 1
            if progress_cb and (done % 50 == 0 or done == total):
                progress_cb(f"DXF {dxf_no} geprüft ({done}/{total})")

        # Обрезаем хвост выше последнего файла, который реально существует
        last_real = max(
            (no for no, rec in result.items() if rec.has_main_dwg),
            default=0
        )

        trimmed = {no: rec for no, rec in result.items() if no <= last_real}

        self._save_files_index_json(trimmed)
        self._files_index = trimmed
        return len(trimmed)

    # ------------------------------------------------------------------
    # Поиск
    # ------------------------------------------------------------------

    def search_by_k_num(self, k_num: str) -> list[DXFSearchResult]:
        """
        Поиск всех DXF-записей для заданного K-номера.
        Объединяет данные из Excel (_excel_index) и файлов (_files_index).
        O(K) где K — количество строк Excel для данного K-номера.
        """
        k_norm = self._normalize_k_num(k_num)
        if not k_norm:
            return []

        excel_rows = self._excel_index.get(k_norm, [])
        results: list[DXFSearchResult] = []

        for row in excel_rows:
            file_rec = self._files_index.get(row.dxf_no)
            results.append(DXFSearchResult(
                dxf_no=row.dxf_no,
                k_num=row.k_num,
                wst=row.wst,
                dicke_mm=row.dicke_mm,
                a_kn_brutto_qm=row.a_kn_brutto_qm,
                laenge_zuschnitt_mm=row.laenge_zuschnitt_mm,
                preis_pro_laenge_eur=row.preis_pro_laenge_eur,
                main_dwg_path=file_rec.main_dwg_path if file_rec else "",
                has_main_dwg=file_rec.has_main_dwg  if file_rec else False,
            ))

        results.sort(key=lambda x: x.dxf_no)
        return results

    def search_by_dxf_no(self, query: str) -> list[DXFSearchResult]:
        """
        Точный поиск по самому DXF-номеру.

        Использует обратный индекс _by_dxf_no — O(1) вместо O(N).
        Если строк в Excel нет, но файл известен — возвращает минимальную запись.

        Важно:
        - этот метод выполняет только точный поиск;
        - для частичного поиска по началу номера используется
          отдельный метод search_by_dxf_partial().
        """
        raw = str(query).strip()
        if not raw or not raw.isdigit():
            return []

        dxf_no = int(raw)
        results: list[DXFSearchResult] = []

        # Быстрый поиск через обратный индекс
        for row in self._by_dxf_no.get(dxf_no, []):
            file_rec = self._files_index.get(dxf_no)
            results.append(DXFSearchResult(
                dxf_no=row.dxf_no,
                k_num=row.k_num,
                wst=row.wst,
                dicke_mm=row.dicke_mm,
                a_kn_brutto_qm=row.a_kn_brutto_qm,
                laenge_zuschnitt_mm=row.laenge_zuschnitt_mm,
                preis_pro_laenge_eur=row.preis_pro_laenge_eur,
                main_dwg_path=file_rec.main_dwg_path if file_rec else "",
                has_main_dwg=file_rec.has_main_dwg if file_rec else False,
            ))

        # Если в Excel ничего нет, но файл известен — минимальная запись
        if not results:
            file_rec = self._files_index.get(dxf_no)
            if file_rec:
                results.append(DXFSearchResult(
                    dxf_no=dxf_no,
                    k_num="",
                    wst="",
                    dicke_mm=None,
                    a_kn_brutto_qm=None,
                    laenge_zuschnitt_mm=None,
                    preis_pro_laenge_eur=None,
                    main_dwg_path=file_rec.main_dwg_path,
                    has_main_dwg=file_rec.has_main_dwg,
                ))

        results.sort(key=lambda x: (x.dxf_no, x.k_num, x.main_dwg_path))
        return results

    def search_by_dxf_partial(self, query: str) -> list[DXFSearchResult]:
        """
        Частичный поиск по началу DXF-номера.

        Правила:
        - допустимы только цифры;
        - выполняется поиск по префиксу строкового представления номера:
          query='116' найдёт 116, 1160, 11601, 11699 и т.п.;
        - результаты собираются как из Excel-индекса (_by_dxf_no),
          так и из файлового индекса (_files_index);
        - если по одному DXF-номеру есть несколько Excel-строк
          (например, для нескольких K-номеров), все они попадают в результат;
        - если Excel-строк нет, но основной DWG-файл известен,
          создаётся минимальная запись.

        Метод предназначен именно для табличного вывода множества совпадений.
        """
        raw = str(query).strip()
        if not raw or not raw.isdigit():
            return []

        results: list[DXFSearchResult] = []
        matched_numbers = sorted(
            no for no in set(self._by_dxf_no.keys()) | set(self._files_index.keys())
            if str(no).startswith(raw)
        )

        for dxf_no in matched_numbers:
            excel_rows = self._by_dxf_no.get(dxf_no, [])
            file_rec = self._files_index.get(dxf_no)

            if excel_rows:
                for row in excel_rows:
                    results.append(DXFSearchResult(
                        dxf_no=row.dxf_no,
                        k_num=row.k_num,
                        wst=row.wst,
                        dicke_mm=row.dicke_mm,
                        a_kn_brutto_qm=row.a_kn_brutto_qm,
                        laenge_zuschnitt_mm=row.laenge_zuschnitt_mm,
                        preis_pro_laenge_eur=row.preis_pro_laenge_eur,
                        main_dwg_path=file_rec.main_dwg_path if file_rec else "",
                        has_main_dwg=file_rec.has_main_dwg if file_rec else False,
                    ))
            elif file_rec:
                results.append(DXFSearchResult(
                    dxf_no=dxf_no,
                    k_num="",
                    wst="",
                    dicke_mm=None,
                    a_kn_brutto_qm=None,
                    laenge_zuschnitt_mm=None,
                    preis_pro_laenge_eur=None,
                    main_dwg_path=file_rec.main_dwg_path,
                    has_main_dwg=file_rec.has_main_dwg,
                ))

        results.sort(key=lambda x: (x.dxf_no, x.k_num, x.main_dwg_path))
        return results

    def ensure_minimum_indexes(self) -> None:
        """
        Создаёт все отсутствующие индексы при первом запуске.
        Вызывается из KFinderFrame.__init__ в фоновом потоке.
        """
        if not self._ranges:
            self.rebuild_ranges_from_workbook()
        if not self._excel_index:
            self.rebuild_excel_index()
        if not self._files_index:
            self.rebuild_files_index_full()


class AppNrRepository:
    """
    Репозиторий серийных номеров аппаратов (Apparate-Nr.).

    Источник данных:
      Excel-файл Apparate-Nr.xlsx.
      Столбец A: серийный номер (например "1234", "1234.1-2").
      Столбец E: K-номер заказа.

    Индексы (хранятся в памяти):
      _by_prefix: str → list[AppNrRecord]
          Быстрый поиск по числовому префиксу.

      _by_k_code: str → list[str]
          Список серийных номеров для K-кода.

      _all_records: list[AppNrRecord]
          Полный список записей для частичного поиска по началу
          полного серийного номера.

    Числовой префикс — часть серийного номера до первой точки:
      "1234"     → "1234"
      "1234.1"   → "1234"
      "1234.1-2" → "1234"
    """

    FULL_K_RE = re.compile(r"^K\d{5}$", re.IGNORECASE)

    def __init__(self, cfg: AppNrConfig = APPNR_CONFIG):
        self.cfg = cfg
        self._by_prefix: dict[str, list[AppNrRecord]] = {}
        self._by_k_code: dict[str, list[str]] = {}
        self._all_records: list[AppNrRecord] = []
        self.reload()

    # ------------------------------------------------------------------
    # Вспомогательные методы
    # ------------------------------------------------------------------

    def _normalize_k_code(self, value) -> str:
        """Приводит значение ячейки к формату K12345."""
        if value is None:
            return ""
        s = str(value).strip().upper()
        if not s:
            return ""
        if re.fullmatch(r"\d{5}", s):
            return f"K{s}"
        if self.FULL_K_RE.fullmatch(s):
            return s
        return ""

    def _normalize_serial_text(self, value) -> str:
        """Нормализует серийный номер: убирает пробелы по краям."""
        if value is None:
            return ""
        return str(value).strip()

    def _extract_prefix(self, serial_text: str) -> str:
        """
        Извлекает числовой префикс серийного номера.
        "1234.1-2" → "1234", "1234" → "1234", "" → "".
        """
        s = serial_text.strip()
        if not s:
            return ""
        head = s.split(".", 1)[0].strip()
        m = re.match(r"^(\d+)", head)
        return m.group(1) if m else ""

    def is_excel_available(self) -> bool:
        return self.cfg.excel_file.exists() and self.cfg.excel_file.is_file()

    # ------------------------------------------------------------------
    # JSON
    # ------------------------------------------------------------------

    def _load_raw(self) -> dict:
        if not self.cfg.index_json.is_file():
            return {"items": []}
        try:
            with self.cfg.index_json.open("r", encoding="utf-8") as f:
                data = json.load(f)
            if not isinstance(data, dict) or "items" not in data:
                return {"items": []}
            return data
        except (OSError, json.JSONDecodeError) as e:
            logger.error(f"AppNrRepository._load_raw: {e}")
            return {"items": []}

    def _save_raw(self, data: dict) -> None:
        self.cfg.index_json.parent.mkdir(parents=True, exist_ok=True)
        with self.cfg.index_json.open("w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)

    def reload(self) -> None:
        """
        Загружает индекс с диска в память.

        Дополнительно формирует:
        - _all_records для частичного поиска по полному serial_no;
        - _by_prefix для быстрого поиска по числовому префиксу;
        - _by_k_code для вывода App.Nr. в статусе K-поиска.
        """
        data = self._load_raw()
        self._by_prefix = {}
        self._by_k_code = {}
        self._all_records = []

        for item in data.get("items", []):
            try:
                rec = AppNrRecord.from_dict(item)
            except (KeyError, TypeError, ValueError):
                continue

            self._all_records.append(rec)
            self._by_prefix.setdefault(rec.serial_prefix, []).append(rec)
            self._by_k_code.setdefault(rec.k_code, []).append(rec.serial_no)

        # Сортировка нужна для стабильного порядка в таблицах и статусе
        self._all_records.sort(key=lambda x: (x.serial_no, x.k_code))

        for prefix in self._by_prefix:
            self._by_prefix[prefix].sort(key=lambda x: (x.serial_no, x.k_code))

        for k_code in self._by_k_code:
            self._by_k_code[k_code] = sorted(set(self._by_k_code[k_code]))

    # ------------------------------------------------------------------
    # Построение индекса
    # ------------------------------------------------------------------

    def rebuild_index(self) -> int:
        """
        Полная перестройка appnr_index.json из Excel-файла.
        Читает столбцы A (серийный номер) и E (K-номер).
        Возвращает количество считанных записей.
        """
        if not self.is_excel_available():
            raise FileNotFoundError(f"Excel-Datei nicht gefunden: {self.cfg.excel_file}")

        wb = load_workbook(self.cfg.excel_file, data_only=True, read_only=True)
        ws = wb.active
        items: list[AppNrRecord] = []
        first = True

        for row in ws.iter_rows(values_only=True):
            if first:
                first = False
                continue

            serial_no = self._normalize_serial_text(row[0] if len(row) > 0 else None)
            k_code = self._normalize_k_code(row[4] if len(row) > 4 else None)

            if not serial_no or not k_code:
                continue

            prefix = self._extract_prefix(serial_no)
            if not prefix:
                continue

            items.append(AppNrRecord(
                serial_no=serial_no,
                serial_prefix=prefix,
                k_code=k_code,
            ))

        wb.close()
        self._save_raw({"items": [x.to_dict() for x in items]})
        self.reload()
        return len(items)

    def ensure_index(self) -> None:
        """Строит индекс если он пуст и Excel-файл доступен."""
        if not self._all_records and self.is_excel_available():
            self.rebuild_index()

    # ------------------------------------------------------------------
    # Поиск
    # ------------------------------------------------------------------

    def search_by_serial(self, query: str) -> list[AppNrRecord]:
        """
        Частичный поиск по началу полного серийного номера.

        Примеры:
        - '1234'     → 1234, 1234.1, 1234.1-2
        - '1234.1'   → 1234.1, 1234.1-2
        - '1234.1-'  → более узкий набор

        Алгоритм:
        1. Нормализуем запрос.
        2. Если запрос пуст — возвращаем [].
        3. Сначала пытаемся использовать быстрый индекс _by_prefix,
           если запрос равен чистому числовому префиксу.
        4. Затем фильтруем по startswith() по полному serial_no.
           Это обеспечивает и совместимость со старым поведением,
           и более точный частичный поиск.
        """
        self.ensure_index()

        q = str(query).strip()
        if not q:
            return []

        # Быстрый путь для чисто числового префикса
        if q.isdigit():
            candidates = self._by_prefix.get(q, [])
            return list(candidates)

        # Общий путь: поиск по началу полного серийного номера
        q_upper = q.upper()
        results = [
            rec for rec in self._all_records
            if rec.serial_no.upper().startswith(q_upper)
        ]
        results.sort(key=lambda x: (x.serial_no, x.k_code))
        return results

    def get_serials_for_k(self, k_code: str) -> list[str]:
        """Возвращает все серийные номера для данного K-кода."""
        self.ensure_index()
        return list(self._by_k_code.get(k_code.strip().upper(), []))


# ============================================================
# GUI-УТИЛИТЫ
# ============================================================

# Цветовая схема в стиле AT-CAD
CLR_BG          = "#508050"   # фон главного окна
CLR_BTN_PRIMARY = "#2980b9"   # синий — основное действие
CLR_BTN_OK      = "#27ae60"   # зелёный — открыть/принять
CLR_BTN_WARN    = "#e67e22"   # оранжевый — осторожное действие
CLR_BTN_DANGER  = "#c0392b"   # красный — закрыть/удалить
CLR_BTN_DARK    = "#2c3e50"   # тёмный — нейтральное действие
CLR_TEXT        = "#ffffff"   # белый текст на кнопках
CLR_LABEL       = "#e8f5e9"   # светло-зелёный для меток
CLR_STATUS_OK   = "#ffffff"   # белый — нормальный статус
CLR_STATUS_WARN = "#f9ca74"   # жёлтый — предупреждение
CLR_STATUS_ERR  = "#f08080"   # красный — ошибка
CLR_BOX_FG      = "#c8e6c9"   # цвет заголовка StaticBox
CLR_PLACEHOLDER = "#A0A0A0"   # серый — placeholder в поле ввода
CLR_INPUT_TEXT  = "#1a3a1a"   # тёмно-зелёный — введённый текст


def _make_gen_button(
    parent: wx.Window,
    label: str,
    color: str,
    size: wx.Size = wx.Size(-1, 36),
    font_size: int = 11,
) -> GenButton:
    """
    Создаёт стилизованную GenButton в стиле AT-CAD.

    Особенности:
    - Hover-эффект: чуть светлее при наведении, темнее при нажатии.
    - Все обработчики мыши вызывают e.Skip(), чтобы EVT_BUTTON срабатывал.
    - Белый жирный шрифт на цветном фоне.
    """
    btn = GenButton(parent, label=label, size=size)
    btn.SetBackgroundColour(wx.Colour(color))
    btn.SetForegroundColour(wx.Colour(CLR_TEXT))
    btn.SetFont(wx.Font(font_size, wx.FONTFAMILY_DEFAULT,
                        wx.FONTSTYLE_NORMAL, wx.FONTWEIGHT_BOLD))
    btn.SetBezelWidth(1)
    btn.SetUseFocusIndicator(False)

    _orig = color

    def _darken(c: str, f: float = 0.88) -> str:
        col = wx.Colour(c)
        return "#{:02x}{:02x}{:02x}".format(
            int(col.Red() * f), int(col.Green() * f), int(col.Blue() * f))

    def _lighten(c: str, f: float = 1.08) -> str:
        col = wx.Colour(c)
        return "#{:02x}{:02x}{:02x}".format(
            min(255, int(col.Red() * f)),
            min(255, int(col.Green() * f)),
            min(255, int(col.Blue() * f)))

    hover = _lighten(_orig)
    press = _darken(_orig)

    # e.Skip() обязателен: без него EVT_BUTTON не сработает
    def on_enter(e): btn.SetBackgroundColour(wx.Colour(hover)); btn.Refresh(); e.Skip()
    def on_leave(e): btn.SetBackgroundColour(wx.Colour(_orig)); btn.Refresh(); e.Skip()
    def on_down(e):  btn.SetBackgroundColour(wx.Colour(press)); btn.Refresh(); e.Skip()
    def on_up(e):    btn.SetBackgroundColour(wx.Colour(_orig)); btn.Refresh(); e.Skip()

    btn.Bind(wx.EVT_ENTER_WINDOW, on_enter)
    btn.Bind(wx.EVT_LEAVE_WINDOW, on_leave)
    btn.Bind(wx.EVT_LEFT_DOWN,    on_down)
    btn.Bind(wx.EVT_LEFT_UP,      on_up)
    return btn


def _static_box_sizer(parent: wx.Window, label: str) -> wx.StaticBoxSizer:
    """
    Создаёт StaticBoxSizer в стиле AT-CAD:
    светло-зелёный жирный заголовок секции.
    """
    box = wx.StaticBox(parent, label=f"  {label}  ")
    box.SetForegroundColour(wx.Colour(CLR_BOX_FG))
    box.SetFont(wx.Font(10, wx.FONTFAMILY_DEFAULT,
                        wx.FONTSTYLE_NORMAL, wx.FONTWEIGHT_BOLD))
    return wx.StaticBoxSizer(box, wx.VERTICAL)


def _open_path(path: Path) -> None:
    """Открывает папку или файл через стандартный обработчик Windows."""
    os.startfile(str(path))


def _show_in_explorer(path: Path) -> None:
    """Открывает папку в Проводнике с выделением файла."""
    subprocess.run(["explorer", "/select,", str(path)], check=False)


# ============================================================
# ДИАЛОГ РЕЗУЛЬТАТОВ K-ПОИСКА
# ============================================================

class ResultsDialog(wx.Dialog):
    """
    Модальное окно с таблицей найденных папок заказов.

    Колонки: K-код | Год | Наличие папки | Путь к папке | Скетчи | DWG.
    Строки без папки (has_folder=False) выделены жёлтым.
    Двойной клик или кнопка «Ordner öffnen» открывают папку заказа.
    """

    def __init__(self, parent: wx.Window, entries: list[KEntry], title: str):
        super().__init__(
            parent,
            title=title or TXT["results_title"],
            style=wx.DEFAULT_DIALOG_STYLE | wx.RESIZE_BORDER,
            size=wx.Size(1100, 560),
        )
        self.entries = entries
        self._build()
        self.CentreOnScreen()

    def _build(self) -> None:
        self.SetBackgroundColour(wx.Colour(CLR_BG))
        outer = wx.BoxSizer(wx.VERTICAL)

        # --- Таблица ---
        tbl_sizer = _static_box_sizer(self, TXT["results_box"])

        self.tree = wx.ListCtrl(
            self,
            style=wx.LC_REPORT | wx.LC_SINGLE_SEL | wx.BORDER_SUNKEN,
        )
        self.tree.SetBackgroundColour(wx.Colour("#f0f4f0"))
        self.tree.SetFont(wx.Font(10, wx.FONTFAMILY_DEFAULT,
                                  wx.FONTSTYLE_NORMAL, wx.FONTWEIGHT_NORMAL))

        cols = [
            (TXT["table_order"],        90),
            (TXT["table_year"],         60),
            (TXT["table_folder_exists"], 130),
            (TXT["table_folder"],       330),
            (TXT["table_sketch"],       280),
            (TXT["table_dwg"],          220),
        ]
        for idx, (hdr, width) in enumerate(cols):
            self.tree.InsertColumn(idx, hdr, width=width)

        for entry in self.entries:
            row = [
                entry.k_code,
                str(entry.year) if entry.year else "—",
                TXT["table_has_folder_yes"] if entry.has_folder else TXT["table_has_folder_no"],
                entry.folder_path if entry.has_folder else "—",
                entry.sketch_path if entry.has_folder else "—",
                entry.dwg_path    if entry.has_folder else "—",
            ]
            idx = self.tree.InsertItem(self.tree.GetItemCount(), row[0])
            for col, val in enumerate(row[1:], 1):
                self.tree.SetItem(idx, col, val)
            if not entry.has_folder:
                self.tree.SetItemBackgroundColour(idx, wx.Colour("#fff0a0"))

        self.tree.Bind(wx.EVT_LIST_ITEM_ACTIVATED, lambda _: self._open_folder())
        tbl_sizer.Add(self.tree, 1, wx.EXPAND | wx.ALL, 6)

        # --- Кнопки ---
        act_sizer = _static_box_sizer(self, TXT["actions_title"])
        btn_row   = wx.BoxSizer(wx.HORIZONTAL)

        actions = [
            (TXT["open_folder"],  CLR_BTN_OK,      self._open_folder),
            (TXT["open_sketch"],  CLR_BTN_OK,      self._open_sketch),
            (TXT["open_dwg"],     CLR_BTN_PRIMARY, self._open_dwg),
            (TXT["close_dialog"], CLR_BTN_DANGER,  self.Close),
        ]
        for label, color, handler in actions:
            btn = _make_gen_button(self, label, color, wx.Size(-1, 34), 10)
            btn.Bind(wx.EVT_BUTTON, lambda _, h=handler: h())
            btn_row.Add(btn, 1, wx.RIGHT, 6)

        act_sizer.Add(btn_row, 0, wx.EXPAND | wx.ALL, 6)
        outer.Add(tbl_sizer, 1, wx.EXPAND | wx.ALL, 10)
        outer.Add(act_sizer, 0, wx.EXPAND | wx.LEFT | wx.RIGHT | wx.BOTTOM, 10)
        self.SetSizer(outer)

    def _selected(self) -> Optional[KEntry]:
        """Возвращает выбранную запись или None."""
        idx = self.tree.GetFirstSelected()
        if idx == wx.NOT_FOUND:
            return None
        code = self.tree.GetItemText(idx, 0)
        return next((e for e in self.entries if e.k_code == code), None)

    def _require_folder(self) -> Optional[KEntry]:
        """
        Возвращает выбранную запись если она имеет папку.
        Показывает MessageBox если ничего не выбрано или папки нет.
        """
        entry = self._selected()
        if not entry:
            wx.MessageBox(TXT["msg_select_entry"], TXT["msg_hint"],
                          wx.OK | wx.ICON_INFORMATION)
            return None
        if not entry.has_folder:
            wx.MessageBox(
                TXT["msg_no_folder_single"].format(code=entry.k_code),
                TXT["msg_no_folder_title"],
                wx.OK | wx.ICON_WARNING,
            )
            return None
        return entry

    def _open_folder(self) -> None:
        entry = self._require_folder()
        if not entry:
            return
        p = Path(entry.folder_path)
        if p.exists():
            _open_path(p)
        else:
            wx.MessageBox(TXT["msg_folder_missing"].format(path=p),
                          TXT["msg_error"], wx.OK | wx.ICON_WARNING)

    def _open_sketch(self) -> None:
        entry = self._require_folder()
        if not entry:
            return
        p = Path(entry.sketch_path)
        if p.exists():
            _open_path(p)
        else:
            wx.MessageBox(TXT["msg_folder_missing"].format(path=p),
                          TXT["msg_error"], wx.OK | wx.ICON_WARNING)

    def _open_dwg(self) -> None:
        entry = self._require_folder()
        if not entry:
            return
        p = Path(entry.dwg_path)
        if p.exists():
            _open_path(p)
        else:
            wx.MessageBox(TXT["msg_file_missing"].format(path=p),
                          TXT["msg_error"], wx.OK | wx.ICON_WARNING)


# ============================================================
# ДИАЛОГ РЕЗУЛЬТАТОВ DXF-ПОИСКА
# ============================================================

class DXFResultsDialog(wx.Dialog):
    """
    Модальное окно с результатами поиска DXF/DWG по K-номеру,
    точному DXF-номеру или частичному совпадению по началу DXF-номера.

    Колонки: DXF | K-Nr. | Werkstoff | Dicke | A Kn brutto | Länge | Preis | DWG.
    Строки без основного DWG-файла (has_main_dwg=False) выделены жёлтым.
    Поддерживает сохранение результатов в TXT и CSV.

    Важно:
    - выбор строки выполняется по индексу строки в таблице, а не только по dxf_no;
    - это необходимо, потому что один и тот же DXF-номер может встречаться
      в нескольких строках для разных K-номеров.
    """

    def __init__(self, parent: wx.Window, results: list[DXFSearchResult], title: str):
        super().__init__(
            parent,
            title=title or TXT["dxf_results_title"],
            style=wx.DEFAULT_DIALOG_STYLE | wx.RESIZE_BORDER,
            size=wx.Size(1120, 560),
        )
        self.results = results
        self._build()
        self.CentreOnScreen()

    def _build(self) -> None:
        self.SetBackgroundColour(wx.Colour(CLR_BG))
        outer = wx.BoxSizer(wx.VERTICAL)

        # --- Таблица результатов ---
        data_sizer = _static_box_sizer(self, TXT["dxf_box"])

        self.tree = wx.ListCtrl(
            self,
            style=wx.LC_REPORT | wx.LC_SINGLE_SEL | wx.BORDER_SUNKEN,
        )
        self.tree.SetBackgroundColour(wx.Colour("#f0f4f0"))
        self.tree.SetFont(wx.Font(
            10,
            wx.FONTFAMILY_DEFAULT,
            wx.FONTSTYLE_NORMAL,
            wx.FONTWEIGHT_NORMAL,
        ))

        cols = [
            (TXT["dxf_col_no"],      90),
            (TXT["dxf_col_k"],       90),
            (TXT["dxf_col_wst"],    110),
            (TXT["dxf_col_dicke"],   90),
            (TXT["dxf_col_area"],   120),
            (TXT["dxf_col_length"], 150),
            (TXT["dxf_col_price"],  100),
            (TXT["dxf_col_file"],   250),
        ]
        for idx, (hdr, width) in enumerate(cols):
            self.tree.InsertColumn(idx, hdr, width=width)

        for res in self.results:
            file_name = Path(res.main_dwg_path).name if res.main_dwg_path else "—"
            row = [
                str(res.dxf_no),
                res.k_num,
                res.wst,
                "" if res.dicke_mm is None else f"{res.dicke_mm:g}",
                "" if res.a_kn_brutto_qm is None else f"{res.a_kn_brutto_qm:g}",
                "" if res.laenge_zuschnitt_mm is None else f"{res.laenge_zuschnitt_mm:g}",
                "" if res.preis_pro_laenge_eur is None else f"{res.preis_pro_laenge_eur:g}",
                file_name,
            ]
            idx = self.tree.InsertItem(self.tree.GetItemCount(), row[0])
            for col, val in enumerate(row[1:], 1):
                self.tree.SetItem(idx, col, val)

            if not res.has_main_dwg:
                self.tree.SetItemBackgroundColour(idx, wx.Colour("#fff0a0"))

        self.tree.Bind(wx.EVT_LIST_ITEM_ACTIVATED, lambda _: self._open_dwg_file())
        data_sizer.Add(self.tree, 1, wx.EXPAND | wx.ALL, 6)

        # --- Кнопки действий ---
        act_sizer = _static_box_sizer(self, TXT["dxf_actions_box"])
        btn_row = wx.BoxSizer(wx.HORIZONTAL)

        actions = [
            (TXT["dxf_open_folder"], CLR_BTN_OK,      self._open_dwg_folder),
            (TXT["dxf_open_file"],   CLR_BTN_PRIMARY, self._open_dwg_file),
            (TXT["dxf_save_file"],   CLR_BTN_DARK,    self._save_to_file),
            (TXT["dxf_close"],       CLR_BTN_DANGER,  self.Close),
        ]
        for label, color, handler in actions:
            btn = _make_gen_button(self, label, color, wx.Size(-1, 34), 10)
            btn.Bind(wx.EVT_BUTTON, lambda _, h=handler: h())
            btn_row.Add(btn, 1, wx.RIGHT, 6)

        act_sizer.Add(btn_row, 0, wx.EXPAND | wx.ALL, 6)
        outer.Add(data_sizer, 1, wx.EXPAND | wx.ALL, 10)
        outer.Add(act_sizer, 0, wx.EXPAND | wx.LEFT | wx.RIGHT | wx.BOTTOM, 10)
        self.SetSizer(outer)

    def _selected(self) -> Optional[DXFSearchResult]:
        """
        Возвращает выбранную строку или None.

        Здесь нельзя искать запись только по dxf_no, потому что один DXF
        может встречаться в нескольких строках. Поэтому берём результат
        напрямую по индексу строки.
        """
        idx = self.tree.GetFirstSelected()
        if idx == wx.NOT_FOUND:
            return None
        if idx < 0 or idx >= len(self.results):
            return None
        return self.results[idx]

    def _require_selected(self) -> Optional[DXFSearchResult]:
        """Возвращает выбранную строку или показывает подсказку."""
        res = self._selected()
        if not res:
            wx.MessageBox(
                TXT["dxf_select_entry"],
                TXT["msg_hint"],
                wx.OK | wx.ICON_INFORMATION,
            )
        return res

    def _open_dwg_folder(self) -> None:
        res = self._require_selected()
        if not res:
            return

        if not res.main_dwg_path:
            wx.MessageBox(
                TXT["dxf_folder_missing"].format(path="—"),
                TXT["msg_error"],
                wx.OK | wx.ICON_WARNING,
            )
            return

        folder = Path(res.main_dwg_path).parent
        if folder.exists():
            _open_path(folder)
        else:
            wx.MessageBox(
                TXT["dxf_folder_missing"].format(path=folder),
                TXT["msg_error"],
                wx.OK | wx.ICON_WARNING,
            )

    def _open_dwg_file(self) -> None:
        res = self._require_selected()
        if not res:
            return

        path = Path(res.main_dwg_path)
        if res.has_main_dwg and path.exists():
            _open_path(path)
        else:
            wx.MessageBox(
                TXT["dxf_file_missing"].format(path=path),
                TXT["msg_error"],
                wx.OK | wx.ICON_WARNING,
            )

    def _save_to_file(self) -> None:
        """Сохраняет результаты в TXT или CSV по выбору пользователя."""
        if not self.results:
            return

        dlg = wx.FileDialog(
            self,
            message=TXT["dxf_save_title"],
            wildcard="Textdateien (*.txt)|*.txt|CSV-Dateien (*.csv)|*.csv",
            style=wx.FD_SAVE | wx.FD_OVERWRITE_PROMPT,
        )
        if dlg.ShowModal() != wx.ID_OK:
            dlg.Destroy()
            return

        target = Path(dlg.GetPath())
        dlg.Destroy()
        is_csv = target.suffix.lower() == ".csv"

        def _fmt(v) -> str:
            return "" if v is None else str(v)

        try:
            if is_csv:
                lines = [
                    "DXF;K-Nr.;Werkstoff;Dicke,mm;A Kn brutto,qm;"
                    "Länge Zuschnitt,mm;Preis/Länge €;DWG"
                ]
                for r in self.results:
                    lines.append(
                        f"{r.dxf_no};{r.k_num};{r.wst};"
                        f"{_fmt(r.dicke_mm)};{_fmt(r.a_kn_brutto_qm)};"
                        f"{_fmt(r.laenge_zuschnitt_mm)};{_fmt(r.preis_pro_laenge_eur)};"
                        f"{r.main_dwg_path}"
                    )
            else:
                lines = []
                sep = "-" * 60
                for r in self.results:
                    lines += [
                        f"DXF: {r.dxf_no}",
                        f"K-Nr.: {r.k_num}",
                        f"Werkstoff: {r.wst}",
                        f"Dicke, mm: {_fmt(r.dicke_mm)}",
                        f"A Kn brutto, qm: {_fmt(r.a_kn_brutto_qm)}",
                        f"Länge Zuschnitt, mm: {_fmt(r.laenge_zuschnitt_mm)}",
                        f"Preis/Länge €: {_fmt(r.preis_pro_laenge_eur)}",
                        f"DWG: {r.main_dwg_path}",
                        sep,
                    ]

            target.write_text("\n".join(lines), encoding="utf-8")
            wx.MessageBox(
                TXT["dxf_save_done"].format(path=target),
                TXT["update_done_title"],
                wx.OK | wx.ICON_INFORMATION,
            )
        except OSError as e:
            wx.MessageBox(str(e), TXT["msg_error"], wx.OK | wx.ICON_ERROR)


class AppNrResultsDialog(wx.Dialog):
    """
    Модальное окно с результатами поиска по Apparate-Nr.

    Показывает:
    - полный серийный номер;
    - числовой префикс;
    - K-номер;
    - наличие папки заказа;
    - путь к папке заказа.

    Поддерживает действия по выбранной строке:
    - открыть папку заказа;
    - открыть папку скетчей;
    - открыть основной DWG;
    - показать связанные DXF/DWG-результаты по K-номеру.

    В окно передаётся родитель KFinderFrame, чтобы можно было
    безопасно использовать уже существующие сервисы поиска K и DXF.
    """

    def __init__(self, parent: "KFinderFrame", records: list[AppNrRecord], title: str):
        super().__init__(
            parent,
            title=title or TXT["app_results_title"],
            style=wx.DEFAULT_DIALOG_STYLE | wx.RESIZE_BORDER,
            size=wx.Size(1100, 560),
        )
        self.owner: KFinderFrame = parent
        self.records = records
        self._build()
        self.CentreOnScreen()

    def _build(self) -> None:
        self.SetBackgroundColour(wx.Colour(CLR_BG))
        outer = wx.BoxSizer(wx.VERTICAL)

        # --- Таблица ---
        tbl_sizer = _static_box_sizer(self, TXT["app_box"])

        self.tree = wx.ListCtrl(
            self,
            style=wx.LC_REPORT | wx.LC_SINGLE_SEL | wx.BORDER_SUNKEN,
        )
        self.tree.SetBackgroundColour(wx.Colour("#f0f4f0"))
        self.tree.SetFont(wx.Font(
            10,
            wx.FONTFAMILY_DEFAULT,
            wx.FONTSTYLE_NORMAL,
            wx.FONTWEIGHT_NORMAL,
        ))

        cols = [
            (TXT["app_col_serial"],        180),
            (TXT["app_col_prefix"],         90),
            (TXT["app_col_k"],              90),
            (TXT["app_col_folder_exists"], 130),
            (TXT["app_col_folder"],        430),
        ]
        for idx, (hdr, width) in enumerate(cols):
            self.tree.InsertColumn(idx, hdr, width=width)

        for rec in self.records:
            entry = self.owner.service.get_or_search(rec.k_code)
            has_folder = bool(entry and entry.has_folder)
            folder_path = entry.folder_path if has_folder else "—"

            row = [
                rec.serial_no,
                rec.serial_prefix,
                rec.k_code,
                TXT["table_has_folder_yes"] if has_folder else TXT["table_has_folder_no"],
                folder_path,
            ]
            idx = self.tree.InsertItem(self.tree.GetItemCount(), row[0])
            for col, val in enumerate(row[1:], 1):
                self.tree.SetItem(idx, col, val)

            if not has_folder:
                self.tree.SetItemBackgroundColour(idx, wx.Colour("#fff0a0"))

        self.tree.Bind(wx.EVT_LIST_ITEM_ACTIVATED, lambda _: self._open_folder())
        tbl_sizer.Add(self.tree, 1, wx.EXPAND | wx.ALL, 6)

        # --- Кнопки ---
        act_sizer = _static_box_sizer(self, TXT["app_actions_box"])
        btn_row = wx.BoxSizer(wx.HORIZONTAL)

        actions = [
            (TXT["open_folder"],  CLR_BTN_OK,      self._open_folder),
            (TXT["open_sketch"],  CLR_BTN_OK,      self._open_sketch),
            (TXT["open_dwg"],     CLR_BTN_PRIMARY, self._open_dwg),
            (TXT["open_dxf"],     CLR_BTN_DARK,    self._show_dxf),
            (TXT["close_dialog"], CLR_BTN_DANGER,  self.Close),
        ]
        for label, color, handler in actions:
            btn = _make_gen_button(self, label, color, wx.Size(-1, 34), 10)
            btn.Bind(wx.EVT_BUTTON, lambda _, h=handler: h())
            btn_row.Add(btn, 1, wx.RIGHT, 6)

        act_sizer.Add(btn_row, 0, wx.EXPAND | wx.ALL, 6)
        outer.Add(tbl_sizer, 1, wx.EXPAND | wx.ALL, 10)
        outer.Add(act_sizer, 0, wx.EXPAND | wx.LEFT | wx.RIGHT | wx.BOTTOM, 10)
        self.SetSizer(outer)

    def _selected_record(self) -> Optional[AppNrRecord]:
        """Возвращает выбранную запись таблицы или None."""
        idx = self.tree.GetFirstSelected()
        if idx == wx.NOT_FOUND:
            return None
        if idx < 0 or idx >= len(self.records):
            return None
        return self.records[idx]

    def _require_selected_record(self) -> Optional[AppNrRecord]:
        """Требует выбранную строку, иначе показывает подсказку."""
        rec = self._selected_record()
        if not rec:
            wx.MessageBox(
                TXT["app_select_entry"],
                TXT["msg_hint"],
                wx.OK | wx.ICON_INFORMATION,
            )
        return rec

    def _require_selected_entry(self) -> Optional[KEntry]:
        """
        Возвращает KEntry для выбранной строки, если у связанного K-кода
        существует папка заказа. Иначе показывает понятное сообщение.
        """
        rec = self._require_selected_record()
        if not rec:
            return None

        entry = self.owner.service.get_or_search(rec.k_code)
        if not entry or not entry.has_folder:
            wx.MessageBox(
                TXT["msg_no_folder_single"].format(code=rec.k_code),
                TXT["msg_no_folder_title"],
                wx.OK | wx.ICON_WARNING,
            )
            return None
        return entry

    def _open_folder(self) -> None:
        entry = self._require_selected_entry()
        if not entry:
            return

        p = Path(entry.folder_path)
        if p.exists():
            _open_path(p)
        else:
            wx.MessageBox(
                TXT["msg_folder_missing"].format(path=p),
                TXT["msg_error"],
                wx.OK | wx.ICON_WARNING,
            )

    def _open_sketch(self) -> None:
        entry = self._require_selected_entry()
        if not entry:
            return

        p = Path(entry.sketch_path)
        if p.exists():
            _open_path(p)
        else:
            wx.MessageBox(
                TXT["msg_folder_missing"].format(path=p),
                TXT["msg_error"],
                wx.OK | wx.ICON_WARNING,
            )

    def _open_dwg(self) -> None:
        entry = self._require_selected_entry()
        if not entry:
            return

        p = Path(entry.dwg_path)
        if p.exists():
            _open_path(p)
        else:
            wx.MessageBox(
                TXT["msg_file_missing"].format(path=p),
                TXT["msg_error"],
                wx.OK | wx.ICON_WARNING,
            )

    def _show_dxf(self) -> None:
        rec = self._require_selected_record()
        if not rec:
            return
        self.owner._show_dxf_results(rec.k_code)

# ============================================================
# СЛУЖЕБНЫЙ ДИАЛОГ
# ============================================================

class ServiceDialog(wx.Dialog):
    """
    Служебное окно с двумя операциями:

    1. Teilaktualisierung
       Быстрое хвостовое обновление:
       - K-индекса заказов;
       - файлового DXF-индекса;
       - App.Nr.-индекса из Excel.

    2. Vollständige Neuindizierung
       Полное перестроение:
       - K-индекса заказов;
       - DXF-диапазонов из листа Key;
       - DXF-Excel-индекса из листа Tabelle1;
       - DXF-файлового индекса;
       - App.Nr.-индекса из Excel.

    Диалог также показывает краткую сводку по текущим индексам.
    """

    def __init__(self, parent: "KFinderFrame", title: str):
        super().__init__(
            parent,
            title=title,
            style=wx.DEFAULT_DIALOG_STYLE,
            size=wx.Size(*parent.cfg.service_window_size),
        )
        self.owner: KFinderFrame = parent
        self._build()
        self._refresh_info()
        self.CentreOnParent()

    def _build(self) -> None:
        self.SetBackgroundColour(wx.Colour(CLR_BG))
        outer = wx.BoxSizer(wx.VERTICAL)

        # --- Информация об индексах ---
        info_sizer = _static_box_sizer(self, TXT["service_info_box"])
        self.info_lbl = wx.StaticText(self, label="—")
        self.info_lbl.SetForegroundColour(wx.Colour(CLR_LABEL))
        self.info_lbl.SetFont(wx.Font(
            9,
            wx.FONTFAMILY_DEFAULT,
            wx.FONTSTYLE_NORMAL,
            wx.FONTWEIGHT_NORMAL,
        ))
        info_sizer.Add(self.info_lbl, 0, wx.ALL | wx.EXPAND, 6)
        outer.Add(info_sizer, 1, wx.EXPAND | wx.ALL, 10)

        # --- Кнопки действий ---
        act_sizer = _static_box_sizer(self, TXT["service_actions_box"])
        btn_row = wx.BoxSizer(wx.HORIZONTAL)

        btn_partial = _make_gen_button(
            self, TXT["service_update_partial"], CLR_BTN_WARN, wx.Size(-1, 36), 10
        )
        btn_full = _make_gen_button(
            self, TXT["service_update_full"], CLR_BTN_PRIMARY, wx.Size(-1, 36), 10
        )
        btn_close = _make_gen_button(
            self, TXT["service_close"], CLR_BTN_DANGER, wx.Size(-1, 36), 10
        )

        btn_partial.Bind(wx.EVT_BUTTON, lambda _: self._run_partial())
        btn_full.Bind(wx.EVT_BUTTON, lambda _: self._run_full())
        btn_close.Bind(wx.EVT_BUTTON, lambda _: self.Close())

        btn_row.Add(btn_partial, 1, wx.RIGHT, 6)
        btn_row.Add(btn_full, 1, wx.RIGHT, 6)
        btn_row.Add(btn_close, 1)

        act_sizer.Add(btn_row, 0, wx.EXPAND | wx.ALL, 6)
        outer.Add(act_sizer, 0, wx.EXPAND | wx.LEFT | wx.RIGHT | wx.BOTTOM, 10)

        self.SetSizer(outer)

    def _refresh_info(self) -> None:
        """
        Обновляет текст сводной информации по индексам.

        Здесь показывается:
        - информация по K-индексу;
        - наличие / количество DXF-диапазонов;
        - количество K-групп в DXF Excel-индексе;
        - количество записей в DXF файловом индексе;
        - количество App.Nr.-записей.
        """
        k_meta = self.owner.index.get_meta()
        k_count = k_meta.get("count", 0)
        k_generated = k_meta.get("generated_at", "—")
        k_root = str(self.owner.cfg.root_dir)

        dxf_ranges_count = len(self.owner.dxf_repo._ranges)
        dxf_excel_groups = len(self.owner.dxf_repo._excel_index)
        dxf_file_count = len(self.owner.dxf_repo._files_index)
        appnr_count = len(self.owner.appnr_repo._all_records)

        lines = [
            "K-Index:",
            f"  Einträge: {k_count}",
            f"  Stand: {k_generated}",
            f"  Root: {k_root}",
            "",
            "DXF-Indexe:",
            f"  Bereiche (Key): {dxf_ranges_count}",
            f"  Excel-Gruppen (nach K-Nr.): {dxf_excel_groups}",
            f"  DWG-Dateien: {dxf_file_count}",
            "",
            "Apparate-Nr.:",
            f"  Einträge: {appnr_count}",
            "",
            "Teilaktualisierung:",
            f"  K: tail_backtrack={self.owner.cfg.tail_backtrack}, "
            f"tail_years={self.owner.cfg.tail_years_to_scan}",
            f"  DXF: file_backtrack={self.owner.dxf_repo.cfg.file_tail_backtrack}, "
            f"file_forward={self.owner.dxf_repo.cfg.file_tail_forward_scan}",
        ]
        self.info_lbl.SetLabel("\n".join(lines))
        self.Layout()

    def _run_partial(self) -> None:
        """
        Запускает частичное обновление всех поддерживаемых индексов.
        После завершения обновляет сводку в окне.
        """
        if wx.MessageBox(
            TXT["service_partial_confirm"],
            TXT["update_confirm_title"],
            wx.YES_NO | wx.ICON_QUESTION,
        ) != wx.YES:
            return

        self.owner.run_partial_update(silent=False, on_done=self._refresh_info)

    def _run_full(self) -> None:
        """
        Запускает полное перестроение всех поддерживаемых индексов.
        После завершения обновляет сводку в окне.
        """
        if wx.MessageBox(
            TXT["service_full_confirm"],
            TXT["msg_rebuild_confirm_title"],
            wx.YES_NO | wx.ICON_QUESTION,
        ) != wx.YES:
            return

        self.owner.run_full_rebuild(on_done=self._refresh_info)


# ============================================================
# ДИАЛОГ «О ПРОГРАММЕ»
# ============================================================

class AboutDialog(wx.Dialog):
    """
    Информационное окно с описанием программы, версией и автором.
    """

    def __init__(self, parent: wx.Window):
        super().__init__(
            parent,
            title=TXT["about_title"],
            style=wx.DEFAULT_DIALOG_STYLE,
            size=wx.Size(430, 410),
        )
        self._build()
        self.CentreOnParent()

    def _build(self) -> None:
        self.SetBackgroundColour(wx.Colour(CLR_BG))
        outer = wx.BoxSizer(wx.VERTICAL)

        # Заголовочный блок (тёмно-зелёный фон)
        header = wx.Panel(self)
        header.SetBackgroundColour(wx.Colour("#2c5f2e"))
        header_sz = wx.BoxSizer(wx.VERTICAL)

        title_lbl = wx.StaticText(header, label=TXT["about_text_title"])
        title_lbl.SetForegroundColour(wx.Colour(CLR_TEXT))
        title_lbl.SetFont(wx.Font(14, wx.FONTFAMILY_DEFAULT,
                                  wx.FONTSTYLE_NORMAL, wx.FONTWEIGHT_BOLD))

        subtitle_lbl = wx.StaticText(header, label=TXT["about_text_subtitle"])
        subtitle_lbl.SetForegroundColour(wx.Colour("#d9f2da"))
        subtitle_lbl.SetFont(wx.Font(12, wx.FONTFAMILY_DEFAULT,
                                     wx.FONTSTYLE_ITALIC, wx.FONTWEIGHT_NORMAL))

        header_sz.Add(title_lbl,    0, wx.ALIGN_CENTER | wx.TOP, 12)
        header_sz.Add(subtitle_lbl, 0, wx.ALIGN_CENTER | wx.TOP | wx.BOTTOM, 10)
        header.SetSizer(header_sz)
        outer.Add(header, 0, wx.EXPAND)

        # Основной текст
        body = wx.Panel(self)
        body.SetBackgroundColour(wx.Colour(CLR_BG))
        body_sz = wx.BoxSizer(wx.VERTICAL)

        body_lbl = wx.StaticText(body, label=TXT["about_text_body"])
        body_lbl.SetForegroundColour(wx.Colour(CLR_LABEL))
        body_lbl.SetFont(wx.Font(12, wx.FONTFAMILY_DEFAULT,
                                 wx.FONTSTYLE_NORMAL, wx.FONTWEIGHT_NORMAL))

        footer_lbl = wx.StaticText(body, label=TXT["about_text_footer"])
        footer_lbl.SetForegroundColour(wx.Colour(CLR_TEXT))
        footer_lbl.SetFont(wx.Font(12, wx.FONTFAMILY_DEFAULT,
                                   wx.FONTSTYLE_NORMAL, wx.FONTWEIGHT_BOLD))

        body_sz.Add(body_lbl,   0, wx.EXPAND | wx.ALL, 16)
        body_sz.Add(footer_lbl, 0, wx.EXPAND | wx.LEFT | wx.RIGHT | wx.BOTTOM, 16)
        body.SetSizer(body_sz)
        outer.Add(body, 1, wx.EXPAND)

        # Кнопка OK
        btn_row = wx.BoxSizer(wx.HORIZONTAL)
        ok_btn  = _make_gen_button(self, TXT["about_ok"], CLR_BTN_PRIMARY, wx.Size(100, 32), 10)
        ok_btn.Bind(wx.EVT_BUTTON, lambda _: self.EndModal(wx.ID_OK))
        btn_row.AddStretchSpacer(1)
        btn_row.Add(ok_btn, 0)

        outer.Add(btn_row, 0, wx.EXPAND | wx.LEFT | wx.RIGHT | wx.BOTTOM, 12)
        self.SetSizer(outer)


# ============================================================
# ГЛАВНОЕ ОКНО
# ============================================================

class KFinderFrame(wx.Frame):
    """
    Главное окно K-Finder.

    Интерфейс:
    - Кнопка выбора режима поиска (K / DXF / App.Nr.) + поле ввода + кнопка очистки.
    - Пять кнопок действий (показать, открыть папку, открыть скетчи, открыть DWG, DXF).
    - Строка статуса (цветовая индикация: белый / жёлтый / красный).
    - Нижняя строка: Service | Info | Выход.

    Режимы поиска:
    - K    — поиск папки заказа по K-номеру.
    - DXF  — поиск DXF/DWG-файлов по DXF-номеру листа.
    - App.Nr. — поиск заказа по серийному номеру аппарата.

    Автообновление:
    При старте (через 300 мс) запускается хвостовое обновление K-индекса в фоне.
    Это позволяет окну открыться мгновенно, а индекс подтянется тихо.
    """

    def __init__(self, cfg: AppConfig = APP_CONFIG):
        super().__init__(
            None,
            title=TXT["app_title"],
            size=wx.Size(*cfg.window_size),
            # Фиксированный размер: окно компактное, изменение размера не нужно
            style=wx.CAPTION | wx.CLOSE_BOX | wx.MINIMIZE_BOX | wx.SYSTEM_MENU,
        )
        self.cfg         = cfg
        self.index       = KIndex(cfg)
        self.service     = SearchService(cfg, self.index)
        self.dxf_repo    = DXFRepository()
        self.appnr_repo  = AppNrRepository()

        # Флаг защиты от двойного запуска индексации
        self._rebuild_running: bool          = False
        # Списки кнопок для блокировки во время индексации
        self._main_buttons:    list[wx.Window] = []
        self._service_buttons: list[wx.Window] = []

        self.SetBackgroundColour(wx.Colour(CLR_BG))
        self.SetMinSize(wx.Size(*cfg.window_size))
        self.SetMaxSize(wx.Size(*cfg.window_size))

        self._center()
        self._build()
        self._set_search_mode(TXT["search_mode_k"], clear_entry=False)
        self._check_root()
        self.Bind(wx.EVT_CLOSE, self._on_close)

        # Автообновление K-индекса при старте (тихое, в фоне)
        if self.cfg.auto_update_on_start:
            wx.CallLater(300, lambda: self._update_tail_async(ask=False, silent=True))

    # ------------------------------------------------------------------
    # Построение UI
    # ------------------------------------------------------------------

    def _center(self) -> None:
        """Центрирует окно на экране."""
        sw, sh = wx.GetDisplaySize()
        w, h   = self.GetSize()
        self.SetPosition(wx.Point((sw - w) // 2, (sh - h) // 2))

    def _build(self) -> None:
        """Строит всё содержимое главного окна."""
        outer      = wx.BoxSizer(wx.VERTICAL)
        ctrl_height = 36
        text_height = 12

        # ── Поле ввода + тип поиска ───────────────────────────────────
        input_sizer = _static_box_sizer(self, TXT["input_box"])
        row         = wx.BoxSizer(wx.HORIZONTAL)

        # Кнопка-переключатель типа поиска (K / DXF / App.Nr.)
        self.search_mode_value = TXT["search_mode_k"]
        self.search_mode_btn   = _make_gen_button(
            self, self.search_mode_value, CLR_BTN_PRIMARY,
            wx.Size(85, ctrl_height), text_height,
        )
        self.search_mode_btn.Bind(wx.EVT_BUTTON, self._on_search_mode_button)

        # Поле ввода с placeholder и обработкой Enter
        self.entry = wx.TextCtrl(
            self,
            style=wx.TE_CENTER | wx.TE_PROCESS_ENTER,
            size=wx.Size(170, ctrl_height),
        )
        self.entry.SetMinSize(wx.Size(170, ctrl_height))
        self.entry.SetFont(wx.Font(18, wx.FONTFAMILY_DEFAULT,
                                   wx.FONTSTYLE_NORMAL, wx.FONTWEIGHT_BOLD))
        self.entry.SetForegroundColour(wx.Colour(CLR_INPUT_TEXT))
        self.entry.Bind(wx.EVT_TEXT_ENTER, lambda _: self._smart_search())
        self.entry.Bind(wx.EVT_SET_FOCUS,  self._on_entry_focus)
        self.entry.Bind(wx.EVT_KILL_FOCUS, self._on_entry_kill_focus)

        # Инициализируем placeholder
        self._current_hint = ""
        self._update_input_hint(force=True)

        # Кнопка очистки поля и сброса режима
        clear_btn = _make_gen_button(self, "✖", CLR_BTN_DANGER,
                                     wx.Size(36, ctrl_height), text_height)
        clear_btn.Bind(wx.EVT_BUTTON, lambda _: self._clear_search_form())

        row.Add(self.search_mode_btn, 0, wx.ALIGN_CENTER_VERTICAL | wx.RIGHT, 8)
        row.Add(self.entry,           1, wx.ALIGN_CENTER_VERTICAL | wx.RIGHT, 6)
        row.Add(clear_btn,            0, wx.ALIGN_CENTER_VERTICAL)
        input_sizer.Add(row, 0, wx.EXPAND | wx.ALL, 6)
        outer.Add(input_sizer, 0, wx.EXPAND | wx.ALL, 10)

        # ── Основные кнопки действий ──────────────────────────────────
        action_sizer = _static_box_sizer(self, TXT["actions_box"])
        actions = [
            (TXT["search_show"], CLR_BTN_PRIMARY, "show"),
            (TXT["open_folder"], CLR_BTN_OK,      "folder"),
            (TXT["open_sketch"], CLR_BTN_OK,      "sketch"),
            (TXT["open_dwg"],    CLR_BTN_PRIMARY, "dwg"),
            (TXT["open_dxf"],    CLR_BTN_DARK,    "dxf"),
        ]
        for label, color, action in actions:
            btn = _make_gen_button(self, label, color, wx.Size(-1, ctrl_height), text_height)
            btn.Bind(wx.EVT_BUTTON, lambda _, a=action: self._handle_action(a))
            action_sizer.Add(btn, 0, wx.EXPAND | wx.LEFT | wx.RIGHT | wx.BOTTOM, 6)
            self._main_buttons.append(btn)
        outer.Add(action_sizer, 0, wx.EXPAND | wx.LEFT | wx.RIGHT, 10)

        # ── Строка статуса ────────────────────────────────────────────
        status_sizer  = _static_box_sizer(self, TXT["status_box"])
        self.status_lbl = wx.StaticText(self, label=TXT["status_ready_short"])
        self.status_lbl.SetForegroundColour(wx.Colour(CLR_STATUS_OK))
        self.status_lbl.SetFont(wx.Font(text_height, wx.FONTFAMILY_DEFAULT,
                                        wx.FONTSTYLE_NORMAL, wx.FONTWEIGHT_BOLD))
        self.status_lbl.Wrap(340)
        status_sizer.Add(self.status_lbl, 0, wx.ALL | wx.EXPAND, 6)
        outer.Add(status_sizer, 0, wx.EXPAND | wx.ALL, 10)

        # ── Нижние кнопки: Service | Info | Выход ─────────────────────
        bottom_row = wx.BoxSizer(wx.HORIZONTAL)

        service_btn = _make_gen_button(
            self, TXT["service_button"], CLR_BTN_WARN, wx.Size(-1, ctrl_height), text_height)
        service_btn.Bind(wx.EVT_BUTTON, lambda _: self._open_service_dialog())

        about_btn = _make_gen_button(
            self, TXT["about_button"], CLR_BTN_DARK, wx.Size(-1, ctrl_height), text_height)
        about_btn.Bind(wx.EVT_BUTTON, lambda _: self._show_about())

        exit_btn = _make_gen_button(
            self, TXT["close_program"], CLR_BTN_DANGER, wx.Size(-1, ctrl_height), text_height)
        exit_btn.Bind(wx.EVT_BUTTON, lambda _: self.Close())

        bottom_row.Add(service_btn, 1, wx.RIGHT, 6)
        bottom_row.Add(about_btn,   1, wx.RIGHT, 6)
        bottom_row.Add(exit_btn,    1)
        outer.Add(bottom_row, 0, wx.EXPAND | wx.LEFT | wx.RIGHT | wx.BOTTOM, 10)

        self.SetSizer(outer)
        self.entry.SetFocus()

    # ------------------------------------------------------------------
    # Placeholder и режим поиска
    # ------------------------------------------------------------------

    def _on_entry_focus(self, event: wx.FocusEvent) -> None:
        """При фокусе убирает placeholder и переключает цвет текста."""
        if self.entry.GetValue() == self._current_hint:
            self.entry.SetValue("")
            self.entry.SetForegroundColour(wx.Colour(CLR_INPUT_TEXT))
        event.Skip()

    def _on_entry_kill_focus(self, event: wx.FocusEvent) -> None:
        """При потере фокуса восстанавливает placeholder если поле пустое."""
        if not self.entry.GetValue().strip():
            self.entry.SetValue(self._current_hint)
            self.entry.SetForegroundColour(wx.Colour(CLR_PLACEHOLDER))
            self.entry.SetInsertionPoint(0)
        event.Skip()

    def _update_input_hint(self, force: bool = False) -> None:
        """
        Обновляет placeholder в зависимости от текущего режима.
        force=True — принудительно устанавливает placeholder в поле,
        если оно пустое или содержит старый placeholder.
        """
        hints = {
            TXT["search_mode_k"]:   TXT["input_hint_k"],
            TXT["search_mode_dxf"]: TXT["input_hint_dxf"],
            TXT["search_mode_app"]: TXT["input_hint_app"],
        }
        new_hint      = hints.get(self.search_mode_value, "")
        current_value = self.entry.GetValue()

        if force:
            if not current_value or current_value == self._current_hint:
                self._current_hint = new_hint
                self.entry.SetValue(new_hint)
                self.entry.SetForegroundColour(wx.Colour(CLR_PLACEHOLDER))
                self.entry.SetInsertionPoint(0)
                return

        self._current_hint = new_hint

    def _get_search_mode(self) -> str:
        return self.search_mode_value

    def _get_search_mode_color(self, mode: str) -> str:
        """Цвет кнопки-переключателя для каждого режима."""
        if mode == TXT["search_mode_dxf"]:
            return CLR_BTN_DARK
        if mode == TXT["search_mode_app"]:
            return CLR_BTN_WARN
        return CLR_BTN_PRIMARY

    def _set_search_mode(self, mode: str, clear_entry: bool = True) -> None:
        """
        Переключает режим поиска: обновляет кнопку, placeholder, фокус.
        clear_entry=True очищает поле ввода (при смене режима вручную).
        """
        self.search_mode_value = mode
        self.search_mode_btn.SetLabel(mode)
        self.search_mode_btn.SetBackgroundColour(wx.Colour(self._get_search_mode_color(mode)))
        self.search_mode_btn.Refresh()
        if clear_entry:
            self.entry.SetValue("")
        self._update_input_hint(force=True)
        self.entry.SetFocus()

    def _on_search_mode_button(self, event: wx.CommandEvent) -> None:
        """Показывает всплывающее меню выбора режима поиска."""
        menu     = wx.Menu()
        item_k   = menu.Append(wx.ID_ANY, TXT["search_mode_k"])
        item_dxf = menu.Append(wx.ID_ANY, TXT["search_mode_dxf"])
        item_app = menu.Append(wx.ID_ANY, TXT["search_mode_app"])

        self.Bind(wx.EVT_MENU, lambda _: self._set_search_mode(TXT["search_mode_k"]),   item_k)
        self.Bind(wx.EVT_MENU, lambda _: self._set_search_mode(TXT["search_mode_dxf"]), item_dxf)
        self.Bind(wx.EVT_MENU, lambda _: self._set_search_mode(TXT["search_mode_app"]), item_app)

        btn = event.GetEventObject()
        if isinstance(btn, wx.Window):
            btn.PopupMenu(menu)
        menu.Destroy()

    def _clear_search_form(self) -> None:
        """Сбрасывает поле ввода и возвращает режим K."""
        self._set_search_mode(TXT["search_mode_k"])

    # ------------------------------------------------------------------
    # Статус и блокировка UI
    # ------------------------------------------------------------------

    def _set_status(self, text: str, level: str = "ok") -> None:
        """
        Обновляет строку статуса.
        level: 'ok' (белый) | 'warn' (жёлтый) | 'err' (красный).
        """
        colors = {
            "ok":   CLR_STATUS_OK,
            "warn": CLR_STATUS_WARN,
            "err":  CLR_STATUS_ERR,
        }
        self.status_lbl.SetLabel(text)
        self.status_lbl.SetForegroundColour(wx.Colour(colors.get(level, CLR_STATUS_OK)))
        self.status_lbl.Wrap(340)
        self.status_lbl.Refresh()
        self.Layout()

    def _set_busy(self, busy: bool) -> None:
        """
        Блокирует или разблокирует UI на время фоновой операции.
        При busy=True включает системный курсор ожидания.
        """
        self.entry.Enable(not busy)
        for btn in self._main_buttons:
            btn.Enable(not busy)
        for btn in self._service_buttons:
            btn.Enable(not busy)

        if busy:
            if not wx.IsBusy():
                wx.BeginBusyCursor()
        else:
            if wx.IsBusy():
                wx.EndBusyCursor()

    # ------------------------------------------------------------------
    # Инициализация
    # ------------------------------------------------------------------

    def _check_root(self) -> None:
        """Предупреждает если корневая папка заказов недоступна."""
        if not self.index.is_root_available():
            self._set_status(TXT["status_root_warn"], "warn")
            wx.MessageBox(
                TXT["msg_root_unavailable"].format(path=self.cfg.root_dir),
                TXT["msg_warning"],
                wx.OK | wx.ICON_WARNING,
            )

    def _on_close(self, _event: wx.CloseEvent) -> None:
        """Уничтожает окно при закрытии (завершает MainLoop)."""
        self.Destroy()

    # ------------------------------------------------------------------
    # Поиск и действия
    # ------------------------------------------------------------------

    def _get_query(self) -> str:
        """Возвращает введённый текст или '' если там placeholder."""
        value = self.entry.GetValue().strip()
        if value == self._current_hint:
            return ""
        return value

    def _get_serials_text_for_k(self, k_code: str) -> str:
        """Возвращает строку серийных номеров для K-кода (для строки статуса)."""
        serials = self.appnr_repo.get_serials_for_k(k_code)
        return ", ".join(serials)

    def _resolve_app_to_k_codes(self, raw: str) -> list[str]:
        """
        Ищет K-коды по введённому серийному номеру или его начальному фрагменту.

        Метод сохранён как вспомогательный для совместимости с остальным кодом.
        Возвращает уникальные K-коды в отсортированном виде.
        """
        records = self.appnr_repo.search_by_serial(raw)
        return sorted({r.k_code for r in records})

    def _normalize_dxf_input(self, raw: str) -> str:
        """
        Нормализует ввод для DXF-режима.

        Допустимы только цифры. Метод принимает как полный DXF-номер,
        так и его начало для частичного поиска.

        Примеры допустимого ввода:
        - 11601   -> точный поиск
        - 116     -> частичный поиск по началу номера
        - 78      -> частичный поиск по началу номера

        Бросает ValueError при любом другом формате.
        """
        s = raw.strip()
        if not s or not s.isdigit():
            raise ValueError(TXT["msg_dxf_input_error"])
        return s

    def _smart_search(self) -> None:
        """
        Обрабатывает нажатие Enter в поле ввода.
        Делегирует поиск в зависимости от текущего режима.
        """
        raw = self._get_query()
        if not raw:
            wx.MessageBox(TXT["msg_input_required"], TXT["msg_hint"],
                          wx.OK | wx.ICON_INFORMATION)
            self.entry.SetFocus()
            return

        mode = self._get_search_mode()

        try:
            if mode == TXT["search_mode_k"]:
                if self.index.is_full_code(raw):
                    self._process_full(self.index.normalize_full(raw), "show")
                else:
                    self._process_partial(raw)
                return

            if mode == TXT["search_mode_dxf"]:
                dxf_no = self._normalize_dxf_input(raw)
                self._process_dxf(dxf_no, "show")
                return

            if mode == TXT["search_mode_app"]:
                self._process_app_nr(raw, "show")
                return

        except ValueError as e:
            wx.MessageBox(str(e), TXT["msg_input_error"], wx.OK | wx.ICON_ERROR)
            self.entry.SetFocus()
            self.entry.SelectAll()

    def _handle_action(self, action: str) -> None:
        """
        Обрабатывает нажатие любой кнопки действия.
        Делегирует в нужный процессор по текущему режиму.
        """
        raw = self._get_query()
        if not raw:
            wx.MessageBox(TXT["msg_input_required"], TXT["msg_hint"],
                          wx.OK | wx.ICON_INFORMATION)
            self.entry.SetFocus()
            return

        mode = self._get_search_mode()

        try:
            if mode == TXT["search_mode_k"]:
                # Кнопка DXF требует полного K-кода
                if action == "dxf":
                    if not self.index.is_full_code(raw):
                        wx.MessageBox(TXT["msg_input_error"], TXT["msg_hint"],
                                      wx.OK | wx.ICON_INFORMATION)
                        self.entry.SetFocus()
                        self.entry.SelectAll()
                        return
                    self._show_dxf_results(self.index.normalize_full(raw))
                    return

                if self.index.is_full_code(raw):
                    self._process_full(self.index.normalize_full(raw), action)
                else:
                    self._process_partial(raw)
                return

            if mode == TXT["search_mode_dxf"]:
                dxf_no = self._normalize_dxf_input(raw)
                self._process_dxf(dxf_no, action)
                return

            if mode == TXT["search_mode_app"]:
                self._process_app_nr(raw, action)
                return

        except ValueError as e:
            wx.MessageBox(str(e), TXT["msg_input_error"], wx.OK | wx.ICON_ERROR)
            self.entry.SetFocus()
            self.entry.SelectAll()
        except (OSError, FileNotFoundError) as e:
            self._set_status(str(e), "err")
            wx.MessageBox(str(e), TXT["msg_error"], wx.OK | wx.ICON_ERROR)

    def _process_dxf(self, dxf_no: str, action: str) -> None:
        """
        Обрабатывает действие в режиме DXF-поиска.

        Логика:
        1. Сначала пытаемся выполнить точный поиск по DXF-номеру.
        2. Если точных совпадений нет — выполняем частичный поиск
           по началу номера.
        3. Для show/dxf всегда показываем таблицу результатов.
        4. Для прямых действий (folder/dwg) открываем первый результат
           только если он найден однозначно точным поиском; иначе тоже
           открываем таблицу, чтобы пользователь сам выбрал строку.

        Такой подход сохраняет старое ожидаемое поведение для полного номера
        и добавляет удобный partial-match без отдельного режима.
        """
        exact_results = self.dxf_repo.search_by_dxf_no(dxf_no)
        partial_results = [] if exact_results else self.dxf_repo.search_by_dxf_partial(dxf_no)

        results = exact_results or partial_results

        if not results:
            wx.MessageBox(
                TXT["msg_dxf_not_found"].format(query=dxf_no),
                TXT["msg_dxf_not_found_title"],
                wx.OK | wx.ICON_INFORMATION,
            )
            self.entry.SetFocus()
            self.entry.SelectAll()
            return

        # Для показа всегда открываем таблицу.
        if action in ("show", "dxf"):
            dlg = DXFResultsDialog(self, results, f"{TXT['dxf_results_title']} — {dxf_no}")
            dlg.ShowModal()
            dlg.Destroy()
            return

        # Для прямых действий безопасно автоматически открывать только
        # результат точного поиска. Для частичного поиска нужна явная
        # выборка строки пользователем.
        if not exact_results:
            dlg = DXFResultsDialog(self, results, f"{TXT['dxf_results_title']} — {dxf_no}")
            dlg.ShowModal()
            dlg.Destroy()
            return

        first = exact_results[0]
        path = Path(first.main_dwg_path)

        if action == "folder":
            folder = path.parent
            if folder.exists():
                _open_path(folder)
            else:
                wx.MessageBox(
                    TXT["dxf_folder_missing"].format(path=folder),
                    TXT["msg_error"],
                    wx.OK | wx.ICON_WARNING,
                )
            return

        if action == "dwg":
            if first.has_main_dwg and path.exists():
                _open_path(path)
            else:
                wx.MessageBox(
                    TXT["dxf_file_missing"].format(path=path),
                    TXT["msg_error"],
                    wx.OK | wx.ICON_WARNING,
                )
            return

        # Режим DXF не поддерживает открытие папки скетчей
        if action == "sketch":
            wx.MessageBox(
                TXT["msg_mode_not_supported"],
                TXT["msg_hint"],
                wx.OK | wx.ICON_INFORMATION,
            )

    def _process_app_nr(self, raw: str, action: str) -> None:
        """
        Обрабатывает действие в режиме App.Nr.-поиска.

        Новое поведение:
        - поиск выполняется по началу полного серийного номера;
        - при множественных совпадениях вместо MessageBox открывается
          табличный AppNrResultsDialog;
        - для show всегда открывается таблица, даже если найден один результат;
        - для прямого действия при одном результате выполняется переход
          сразу к связанному K-коду;
        - для прямого действия при нескольких результатах открывается таблица,
          чтобы пользователь сам выбрал нужную строку.
        """
        records = self.appnr_repo.search_by_serial(raw)

        if not records:
            wx.MessageBox(
                TXT["msg_app_not_found"].format(query=raw),
                TXT["msg_app_not_found_title"],
                wx.OK | wx.ICON_INFORMATION,
            )
            self.entry.SetFocus()
            self.entry.SelectAll()
            return

        # Таблица по явному show или при неоднозначности
        if action == "show" or len(records) > 1:
            dlg = AppNrResultsDialog(
                self,
                records,
                f"{TXT['app_results_title']} — {raw}",
            )
            dlg.ShowModal()
            dlg.Destroy()
            return

        # Ниже остаётся только один однозначный результат
        rec = records[0]
        k_code = rec.k_code

        if action == "dxf":
            self._show_dxf_results(k_code)
            return

        self._process_full(k_code, action)

    def _process_full(self, k_code: str, action: str) -> None:
        """
        Обрабатывает полный K-код: поиск → отображение / открытие.
        Показывает серийный номер аппарата в строке статуса если найден.
        """
        self._set_status(TXT["status_searching"].format(code=k_code), "ok")
        entry = self.service.get_or_search(k_code)

        if not entry:
            self._set_status(TXT["status_not_found"].format(code=k_code), "warn")
            wx.MessageBox(TXT["msg_not_found_full"].format(code=k_code),
                          TXT["msg_not_found_title"], wx.OK | wx.ICON_INFORMATION)
            self.entry.SetFocus()
            self.entry.SelectAll()
            return

        if not entry.has_folder:
            self._set_status(TXT["status_no_folder"].format(code=k_code), "warn")
            wx.MessageBox(TXT["msg_no_folder_full"].format(code=k_code),
                          TXT["msg_no_folder_title"], wx.OK | wx.ICON_WARNING)
            self.entry.SetFocus()
            self.entry.SelectAll()
            return

        # Показываем серийный номер аппарата в статусе (если есть)
        serials_text = self._get_serials_text_for_k(k_code)
        if serials_text:
            self._set_status(
                TXT["status_found_with_serial"].format(code=k_code, serials=serials_text), "ok")
        else:
            self._set_status(TXT["status_found"].format(code=k_code), "ok")

        if action == "show":
            title = k_code
            if serials_text:
                short = serials_text if len(serials_text) <= 60 else serials_text[:57] + "..."
                title = f"{k_code} | App. Nr.: {short}"
            dlg = ResultsDialog(self, [entry], title)
            dlg.ShowModal()
            dlg.Destroy()
            return

        path_map = {
            "folder": Path(entry.folder_path),
            "sketch": Path(entry.sketch_path),
            "dwg":    Path(entry.dwg_path),
        }
        path = path_map.get(action)
        if path is None:
            return

        if path.exists():
            _open_path(path)
        else:
            msg = TXT["msg_file_missing"] if action == "dwg" else TXT["msg_folder_missing"]
            wx.MessageBox(msg.format(path=path), TXT["msg_error"], wx.OK | wx.ICON_WARNING)

    def _process_partial(self, raw: str) -> None:
        """Обрабатывает частичный K-запрос: показывает список совпадений."""
        results = self.index.find_partial(raw)

        if not results:
            self._set_status(TXT["status_no_hits"], "warn")
            wx.MessageBox(TXT["msg_no_hits"].format(query=raw),
                          TXT["msg_no_hits_title"], wx.OK | wx.ICON_INFORMATION)
            self.entry.SetFocus()
            self.entry.SelectAll()
            return

        if len(results) == 1:
            entry = results[0]

            if not entry.has_folder:
                self._set_status(TXT["status_no_folder"].format(code=entry.k_code), "warn")
                wx.MessageBox(TXT["msg_no_folder_single"].format(code=entry.k_code),
                              TXT["msg_no_folder_title"], wx.OK | wx.ICON_WARNING)
                self.entry.SetFocus()
                self.entry.SelectAll()
                return

            if self.cfg.auto_open_single:
                p = Path(entry.folder_path)
                if p.exists():
                    _open_path(p)
                    self._set_status(TXT["status_found_one"].format(code=entry.k_code), "ok")
                    return

            if self.cfg.auto_show_single:
                dlg = ResultsDialog(self, [entry], entry.k_code)
                dlg.ShowModal()
                dlg.Destroy()
                self._set_status(TXT["status_found_one"].format(code=entry.k_code), "ok")
                return

        dlg = ResultsDialog(self, results, f"Treffer für '{raw}'")
        dlg.ShowModal()
        dlg.Destroy()
        self._set_status(TXT["status_found_many"].format(count=len(results)), "ok")

    # ------------------------------------------------------------------
    # DXF-результаты
    # ------------------------------------------------------------------

    def _show_dxf_results(self, k_code: str) -> None:
        """Ищет DXF-данные для K-кода и показывает DXFResultsDialog."""
        results = self.dxf_repo.search_by_k_num(k_code)

        if not results:
            wx.MessageBox(TXT["dxf_no_hits"].format(code=k_code),
                          TXT["dxf_no_hits_title"], wx.OK | wx.ICON_INFORMATION)
            self.entry.SetFocus()
            self.entry.SelectAll()
            return

        dlg = DXFResultsDialog(self, results, f"{TXT['dxf_results_title']} — {k_code}")
        dlg.ShowModal()
        dlg.Destroy()

    # ------------------------------------------------------------------
    # Служебные диалоги
    # ------------------------------------------------------------------

    def _open_service_dialog(self) -> None:
        dlg = ServiceDialog(self, TXT["service_dialog_title"])
        dlg.ShowModal()
        dlg.Destroy()

    def _show_about(self) -> None:
        dlg = AboutDialog(self)
        dlg.ShowModal()
        dlg.Destroy()

    # ------------------------------------------------------------------
    # Обновление K-индекса
    # ------------------------------------------------------------------

    def _update_tail_async(
            self,
            ask: bool = True,
            silent: bool = False,
            on_done: Optional[Callable] = None,
    ) -> None:
        """
        Запускает частичное обновление всех индексов в фоновом потоке.

        Что обновляется:
        1. K-индекс заказов:
           - хвостовое обновление последних лет.

        2. DXF-индексы:
           - при необходимости строится карта диапазонов из Excel (лист Key);
           - выполняется хвостовое обновление файлового индекса DWG;
           - Excel-индекс DXF перестраивается полностью, т.к. таблица является
             основным источником и обычно меняется целостно.

        3. App.Nr.:
           - индекс серийных номеров перестраивается полностью из Excel.

        ask=True     — сначала спрашивает подтверждение.
        silent=True  — не показывает сообщение по завершению.
        on_done      — callback, вызывается после завершения через wx.CallAfter.
        """
        if self._rebuild_running:
            if not silent:
                wx.MessageBox(
                    TXT["update_already_running"],
                    TXT["update_confirm_title"],
                    wx.OK | wx.ICON_INFORMATION,
                )
            return

        if ask:
            if wx.MessageBox(
                    TXT["update_confirm"],
                    TXT["update_confirm_title"],
                    wx.YES_NO | wx.ICON_QUESTION,
            ) != wx.YES:
                return

        self._rebuild_running = True
        self._set_busy(True)
        self._set_status(
            TXT["status_start_update"] if silent else TXT["service_update_running"],
            "warn",
        )

        def worker():
            try:
                # ----------------------------------------------------------
                # 1. K-индекс
                # ----------------------------------------------------------
                self.after(lambda: self._set_status("K-Index: Teilaktualisierung…", "warn"))
                k_count = self.index.update_tail(
                    backtrack=self.cfg.tail_backtrack,
                    tail_years_to_scan=self.cfg.tail_years_to_scan,
                    progress_cb=lambda m: self.after(lambda msg=m: self._set_status(msg, "warn")),
                )

                # ----------------------------------------------------------
                # 2. DXF-индексы
                # ----------------------------------------------------------
                self.after(lambda: self._set_status("DXF: Bereiche prüfen…", "warn"))
                if not self.dxf_repo._ranges:
                    self.dxf_repo.rebuild_ranges_from_workbook()

                self.after(lambda: self._set_status("DXF: Excel-Index aktualisieren…", "warn"))
                dxf_excel_count = self.dxf_repo.rebuild_excel_index()

                self.after(lambda: self._set_status("DXF: Datei-Index aktualisieren…", "warn"))
                dxf_file_count = self.dxf_repo.update_files_index_tail(
                    backtrack=self.dxf_repo.cfg.file_tail_backtrack,
                    forward_scan=self.dxf_repo.cfg.file_tail_forward_scan,
                    progress_cb=lambda m: self.after(lambda msg=m: self._set_status(msg, "warn")),
                )

                # ----------------------------------------------------------
                # 3. App.Nr.
                # ----------------------------------------------------------
                self.after(lambda: self._set_status("App.Nr.: Index aktualisieren…", "warn"))
                appnr_count = self.appnr_repo.rebuild_index()

                # Перезагрузка индексов в память для надёжности
                self.after(self.index.reload)
                self.after(self.dxf_repo.reload_all)
                self.after(self.appnr_repo.reload)

                done_msg = (
                    f"K: {k_count} | "
                    f"DXF-Excel: {dxf_excel_count} | "
                    f"DXF-Dateien: {dxf_file_count} | "
                    f"App.Nr.: {appnr_count}"
                )

                self.after(lambda: self._set_status(done_msg, "ok"))

                if not silent:
                    self.after(lambda: wx.MessageBox(
                        "Indexdaten wurden aktualisiert.\n\n"
                        f"K-Index: {k_count}\n"
                        f"DXF-Excel: {dxf_excel_count}\n"
                        f"DXF-Dateien: {dxf_file_count}\n"
                        f"Apparate-Nr.: {appnr_count}",
                        TXT["update_done_title"],
                        wx.OK | wx.ICON_INFORMATION,
                    ))

                if on_done:
                    self.after(on_done)

            except (OSError, FileNotFoundError, KeyError) as e:
                logger.error(f"update_tail_all: {e}")
                self.after(lambda: self._set_status(TXT["service_update_error"], "err"))
                if not silent:
                    self.after(lambda: wx.MessageBox(
                        str(e),
                        TXT["msg_error"],
                        wx.OK | wx.ICON_ERROR,
                    ))
            finally:
                self.after(lambda: setattr(self, "_rebuild_running", False))
                self.after(lambda: self._set_busy(False))

        threading.Thread(target=worker, daemon=True).start()

    def _rebuild_async(
            self,
            ask: bool = True,
            on_done: Optional[Callable] = None,
    ) -> None:
        """
        Запускает полное перестроение всех индексов в фоновом потоке.

        Что перестраивается:
        1. K-индекс заказов — полный rebuild().
        2. DXF-диапазоны — из листа Key.
        3. DXF Excel-индекс — из листа Tabelle1.
        4. DXF файловый индекс — полный проход по диапазону номеров.
        5. App.Nr.-индекс — полное чтение из Excel.

        ask=True — сначала спрашивает подтверждение.
        """
        if self._rebuild_running:
            wx.MessageBox(
                TXT["msg_rebuild_already_running"],
                TXT["msg_rebuild_confirm_title"],
                wx.OK | wx.ICON_INFORMATION,
            )
            return

        if ask:
            if wx.MessageBox(
                    TXT["msg_rebuild_confirm"],
                    TXT["msg_rebuild_confirm_title"],
                    wx.YES_NO | wx.ICON_QUESTION,
            ) != wx.YES:
                return

        self._rebuild_running = True
        self._set_busy(True)
        self._set_status(TXT["service_rebuild_running"], "warn")

        def worker():
            try:
                # ----------------------------------------------------------
                # 1. K-индекс
                # ----------------------------------------------------------
                self.after(lambda: self._set_status("K-Index: Vollaufbau…", "warn"))
                k_count = self.index.rebuild(
                    progress_cb=lambda m: self.after(lambda msg=m: self._set_status(msg, "warn")),
                )

                # ----------------------------------------------------------
                # 2. DXF-индексы
                # ----------------------------------------------------------
                self.after(lambda: self._set_status("DXF: Bereiche aus Excel…", "warn"))
                dxf_ranges_count = self.dxf_repo.rebuild_ranges_from_workbook()

                self.after(lambda: self._set_status("DXF: Excel-Index Vollaufbau…", "warn"))
                dxf_excel_count = self.dxf_repo.rebuild_excel_index()

                self.after(lambda: self._set_status("DXF: Datei-Index Vollaufbau…", "warn"))
                dxf_file_count = self.dxf_repo.rebuild_files_index_full(
                    progress_cb=lambda m: self.after(lambda msg=m: self._set_status(msg, "warn")),
                )

                # ----------------------------------------------------------
                # 3. App.Nr.
                # ----------------------------------------------------------
                self.after(lambda: self._set_status("App.Nr.: Vollaufbau…", "warn"))
                appnr_count = self.appnr_repo.rebuild_index()

                # Перечитать всё с диска после построения
                self.after(self.index.reload)
                self.after(self.dxf_repo.reload_all)
                self.after(self.appnr_repo.reload)

                done_msg = (
                    f"K: {k_count} | "
                    f"DXF-Bereiche: {dxf_ranges_count} | "
                    f"DXF-Excel: {dxf_excel_count} | "
                    f"DXF-Dateien: {dxf_file_count} | "
                    f"App.Nr.: {appnr_count}"
                )
                self.after(lambda: self._set_status(done_msg, "ok"))

                self.after(lambda: wx.MessageBox(
                    "Indexdaten wurden vollständig neu aufgebaut.\n\n"
                    f"K-Index: {k_count}\n"
                    f"DXF-Bereiche: {dxf_ranges_count}\n"
                    f"DXF-Excel: {dxf_excel_count}\n"
                    f"DXF-Dateien: {dxf_file_count}\n"
                    f"Apparate-Nr.: {appnr_count}",
                    TXT["msg_rebuild_done_title"],
                    wx.OK | wx.ICON_INFORMATION,
                ))

                if on_done:
                    self.after(on_done)

            except (OSError, FileNotFoundError, KeyError) as e:
                logger.error(f"rebuild_all: {e}")
                self.after(lambda: self._set_status(TXT["service_rebuild_error"], "err"))
                self.after(lambda: wx.MessageBox(
                    str(e),
                    TXT["msg_error"],
                    wx.OK | wx.ICON_ERROR,
                ))
            finally:
                self.after(lambda: setattr(self, "_rebuild_running", False))
                self.after(lambda: self._set_busy(False))

        threading.Thread(target=worker, daemon=True).start()

    def after(self, func: Callable) -> None:
        """Безопасный вызов функции из фонового потока через wx.CallAfter."""
        wx.CallAfter(func)

    # ------------------------------------------------------------------
    # Публичные методы для ServiceDialog
    # ------------------------------------------------------------------

    def run_partial_update(self, silent: bool = False, on_done: Optional[Callable] = None) -> None:
        """Публичный запуск хвостового обновления (вызывается из ServiceDialog)."""
        self._update_tail_async(ask=False, silent=silent, on_done=on_done)

    def run_full_rebuild(self, on_done: Optional[Callable] = None) -> None:
        """Публичный запуск полного перестроения (вызывается из ServiceDialog)."""
        self._rebuild_async(ask=False, on_done=on_done)


# ============================================================
# ТОЧКА ВХОДА
# ============================================================

def main() -> None:
    app   = wx.App(False)
    frame = KFinderFrame(APP_CONFIG)
    frame.Show()
    app.MainLoop()


if __name__ == "__main__":
    main()