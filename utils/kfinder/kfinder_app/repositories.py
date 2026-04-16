"""
repositories.py
===============

Репозитории данных приложения K-Finder.

Содержит:
1. DXFRepository
   Работа с DXF-данными:
   - Excel-индекс из DXF-2017.xlsm
   - файловый индекс DWG
   - поиск по DXF-номеру
   - поиск по K-номеру
   - rebuild индексов

2. AppNrRepository
   Работа с серийными номерами Apparate-Nr.:
   - полный rebuild из Excel
   - хвостовой rebuild по числовому prefix
   - partial search по началу полного serial_no

Принципы
--------
- GUI-логики здесь нет.
- Все пути берутся из AppSettings.
- Все индексы хранятся в JSON в папке data/.
- Модуль не зависит от main_frame.py.

Важно
-----
DXF Excel-индекс хранится в JSON в структуре:
{
    "K12345": [
        { ... DXFExcelRecord ... },
        { ... DXFExcelRecord ... }
    ],
    ...
}

После загрузки дополнительно строится in-memory индекс:
- _by_k_num   : K-номер -> список DXFExcelRecord
- _by_dxf_no  : DXF-номер -> список DXFExcelRecord

Это позволяет быстро искать и по K, и по DXF.
"""

from __future__ import annotations

import json
import re
from pathlib import Path
from typing import Callable, Optional

from openpyxl import load_workbook

from .config import AppSettings
from .logging_setup import get_logger
from .models import (
    AppNrRecord,
    DXFExcelRecord,
    DXFFileRecord,
    DXFRange,
    DXFSearchResult,
)

logger = get_logger()


# ============================================================
# DXF Repository
# ============================================================

class DXFRepository:
    """
    Репозиторий DXF-данных.

    Источники:
    ----------
    1. Excel-файл DXF-2017.xlsm:
       - лист "Key"      -> диапазоны DXF-номеров
       - лист "Tabelle1" -> технологические данные по DXF

    2. Файловая система:
       - основные DWG-файлы: xxxxx.dwg

    Индексы:
    --------
    _ranges:
        Список DXFRange, полученный из листа Key.

    _excel_index:
        Словарь:
            K-номер -> list[DXFExcelRecord]

    _files_index:
        Словарь:
            dxf_no -> DXFFileRecord

    _by_k_num:
        In-memory alias к _excel_index.

    _by_dxf_no:
        Обратный индекс:
            dxf_no -> list[DXFExcelRecord]
    """

    PRIMARY_DWG_RE = re.compile(r"^(\d+)\.dwg$", re.IGNORECASE)
    FULL_K_RE = re.compile(r"^K\d{5}$", re.IGNORECASE)

    def __init__(self, settings: AppSettings):
        self.settings = settings
        self.paths = settings.paths
        self.indexing = settings.indexing
        self.data_files = settings.data_files

        self._ranges: list[DXFRange] = []
        self._excel_index: dict[str, list[DXFExcelRecord]] = {}
        self._files_index: dict[int, DXFFileRecord] = {}

        self._by_k_num: dict[str, list[DXFExcelRecord]] = {}
        self._by_dxf_no: dict[int, list[DXFExcelRecord]] = {}

        self.reload_all()

    # --------------------------------------------------------
    # Общие утилиты
    # --------------------------------------------------------

    def _ensure_parent(self, path: Path) -> None:
        """
        Создаёт родительскую директорию, если её нет.
        """
        path.parent.mkdir(parents=True, exist_ok=True)

    def _normalize_k_num(self, value) -> str:
        """
        Приводит значение к формату K12345.

        Допустимо:
        - 12345
        - K12345
        """
        if value is None:
            return ""

        s = str(value).strip().upper()
        if not s:
            return ""

        if re.fullmatch(r"\d{5}", s):
            return f"K{s}"

        if self.FULL_K_RE.fullmatch(s):
            return s

        return ""

    def _safe_float(self, value) -> Optional[float]:
        """
        Безопасно приводит значение к float.
        """
        if value is None or value == "":
            return None
        try:
            return float(value)
        except (TypeError, ValueError):
            return None

    def is_dxf_root_available(self) -> bool:
        """
        Проверяет доступность корневой папки DXF-файлов.
        """
        return self.paths.dxf_root_dir.exists() and self.paths.dxf_root_dir.is_dir()

    def is_excel_available(self) -> bool:
        """
        Проверяет доступность Excel-файла DXF.
        """
        return self.paths.dxf_excel_file.exists() and self.paths.dxf_excel_file.is_file()

    def is_primary_dwg_name(self, filename: str) -> bool:
        """
        True, если имя файла является основным DWG:
        только цифры + .dwg, без суффиксов.
        """
        return bool(self.PRIMARY_DWG_RE.fullmatch(filename.strip()))

    def _get_range_for_dxf(self, dxf_no: int) -> Optional[DXFRange]:
        """
        Возвращает диапазон для указанного DXF-номера.
        """
        for item in self._ranges:
            if item.min_no <= dxf_no <= item.max_no:
                return item
        return None

    # --------------------------------------------------------
    # JSON: ranges
    # --------------------------------------------------------

    def _load_ranges_json(self) -> list[DXFRange]:
        path = self.data_files.dxf_ranges_json
        if not path.is_file():
            return []

        try:
            with path.open("r", encoding="utf-8") as f:
                data = json.load(f)
            return [DXFRange.from_dict(item) for item in data]
        except (OSError, json.JSONDecodeError, KeyError, TypeError, ValueError) as e:
            logger.error(f"DXFRepository._load_ranges_json: {e}")
            return []

    def _save_ranges_json(self, ranges: list[DXFRange]) -> None:
        path = self.data_files.dxf_ranges_json
        self._ensure_parent(path)

        with path.open("w", encoding="utf-8") as f:
            json.dump([r.to_dict() for r in ranges], f, ensure_ascii=False, indent=2)

    # --------------------------------------------------------
    # JSON: excel index
    # --------------------------------------------------------

    def _load_excel_index_json(self) -> dict[str, list[DXFExcelRecord]]:
        path = self.data_files.dxf_excel_index_json
        if not path.is_file():
            return {}

        try:
            with path.open("r", encoding="utf-8") as f:
                data = json.load(f)

            return {
                k_num: [DXFExcelRecord.from_dict(item) for item in items]
                for k_num, items in data.items()
            }
        except (OSError, json.JSONDecodeError, KeyError, TypeError, ValueError) as e:
            logger.error(f"DXFRepository._load_excel_index_json: {e}")
            return {}

    def _save_excel_index_json(self, index_data: dict[str, list[DXFExcelRecord]]) -> None:
        path = self.data_files.dxf_excel_index_json
        self._ensure_parent(path)

        payload = {
            k_num: [item.to_dict() for item in items]
            for k_num, items in sorted(index_data.items())
        }

        with path.open("w", encoding="utf-8") as f:
            json.dump(payload, f, ensure_ascii=False, indent=2)

    # --------------------------------------------------------
    # JSON: files index
    # --------------------------------------------------------

    def _load_files_index_json(self) -> dict[int, DXFFileRecord]:
        path = self.data_files.dxf_files_index_json
        if not path.is_file():
            return {}

        try:
            with path.open("r", encoding="utf-8") as f:
                data = json.load(f)

            return {
                int(key): DXFFileRecord.from_dict(item)
                for key, item in data.items()
            }
        except (OSError, json.JSONDecodeError, KeyError, TypeError, ValueError) as e:
            logger.error(f"DXFRepository._load_files_index_json: {e}")
            return {}

    def _save_files_index_json(self, index_data: dict[int, DXFFileRecord]) -> None:
        path = self.data_files.dxf_files_index_json
        self._ensure_parent(path)

        payload = {
            str(dxf_no): item.to_dict()
            for dxf_no, item in sorted(index_data.items())
        }

        with path.open("w", encoding="utf-8") as f:
            json.dump(payload, f, ensure_ascii=False, indent=2)

    # --------------------------------------------------------
    # In-memory rebuild
    # --------------------------------------------------------

    def _rebuild_memory_indexes(self) -> None:
        """
        Перестраивает in-memory индексы:
        - _by_k_num
        - _by_dxf_no
        """
        self._by_k_num = dict(self._excel_index)
        self._by_dxf_no = {}

        for records in self._excel_index.values():
            for rec in records:
                self._by_dxf_no.setdefault(rec.dxf_no, []).append(rec)

    def reload_all(self) -> None:
        """
        Загружает все индексы с диска в память.
        """
        self._ranges = self._load_ranges_json()
        self._excel_index = self._load_excel_index_json()
        self._files_index = self._load_files_index_json()
        self._rebuild_memory_indexes()

    # --------------------------------------------------------
    # Rebuild ranges from workbook
    # --------------------------------------------------------

    def rebuild_ranges_from_workbook(self) -> int:
        """
        Строит dxf_ranges.json из листа 'Key'.

        Формат:
        A = min_no
        B = max_no
        C = folder_name
        """
        if not self.is_excel_available():
            raise FileNotFoundError(f"Excel-Datei nicht gefunden: {self.paths.dxf_excel_file}")

        wb = load_workbook(
            self.paths.dxf_excel_file,
            data_only=True,
            read_only=True,
            keep_vba=True,
        )

        if "Key" not in wb.sheetnames:
            wb.close()
            raise KeyError("Blatt 'Key' wurde in der Excel-Datei nicht gefunden.")

        ws = wb["Key"]
        result: list[DXFRange] = []
        first = True

        for row in ws.iter_rows(values_only=True):
            if first:
                first = False
                continue

            min_no, max_no, folder_name = (row + (None, None, None))[:3]
            if min_no is None or max_no is None or folder_name is None:
                continue

            try:
                result.append(DXFRange(
                    min_no=int(min_no),
                    max_no=int(max_no),
                    folder_name=str(folder_name).strip(),
                ))
            except (TypeError, ValueError):
                continue

        wb.close()

        self._save_ranges_json(result)
        self._ranges = result
        return len(result)

    # --------------------------------------------------------
    # Rebuild excel index
    # --------------------------------------------------------

    def rebuild_excel_index(self) -> int:
        """
        Строит dxf_excel_index.json из листа 'Tabelle1'.

        Используемые колонки:
        A=0   DXF-Nummer
        B=1   K-Nummer
        C=2   Schlüssel
        D=3   Werkstoff
        E=4   Dicke
        F=5   Ch.Nr.
        I=8   A Kn brutto
        K=10  Länge Zuschnitt
        N=13  Preis/Länge
        S=18  Bemerkung
        """
        if not self.is_excel_available():
            raise FileNotFoundError(f"Excel-Datei nicht gefunden: {self.paths.dxf_excel_file}")

        wb = load_workbook(
            self.paths.dxf_excel_file,
            data_only=True,
            read_only=True,
            keep_vba=True,
        )

        if "Tabelle1" not in wb.sheetnames:
            wb.close()
            raise KeyError("Blatt 'Tabelle1' wurde in der Excel-Datei nicht gefunden.")

        ws = wb["Tabelle1"]
        index_data: dict[str, list[DXFExcelRecord]] = {}
        first = True
        total = 0

        for row in ws.iter_rows(values_only=True):
            if first:
                first = False
                continue

            def _col(i: int):
                return row[i] if len(row) > i else None

            dxf_no_raw = _col(0)
            k_raw = _col(1)

            if dxf_no_raw is None or k_raw is None:
                continue

            try:
                dxf_no_int = int(dxf_no_raw)
            except (TypeError, ValueError):
                continue

            k_num = self._normalize_k_num(k_raw)
            if not k_num:
                continue

            rec = DXFExcelRecord(
                dxf_no=dxf_no_int,
                k_num=k_num,
                schluessel=str(_col(2) or ""),
                wst=str(_col(3) or ""),
                dicke_mm=self._safe_float(_col(4)),
                ch_nr=str(_col(5) or ""),
                a_kn_brutto_qm=self._safe_float(_col(8)),
                laenge_zuschnitt_mm=self._safe_float(_col(10)),
                preis_pro_laenge_eur=self._safe_float(_col(13)),
                bemerkung=str(_col(18) or ""),
            )

            index_data.setdefault(k_num, []).append(rec)
            total += 1

        wb.close()

        for k_num in index_data:
            index_data[k_num].sort(key=lambda x: (x.dxf_no, x.schluessel))

        self._save_excel_index_json(index_data)
        self._excel_index = index_data
        self._rebuild_memory_indexes()
        return total

    # --------------------------------------------------------
    # File index rebuild
    # --------------------------------------------------------

    def _build_file_record(self, dxf_no: int) -> DXFFileRecord:
        """
        Строит DXFFileRecord для одного DXF-номера.
        """
        dxfrange = self._get_range_for_dxf(dxf_no)

        if dxfrange is None:
            return DXFFileRecord(
                dxf_no=dxf_no,
                folder_name="",
                folder_path="",
                main_dwg_path="",
                has_main_dwg=False,
            )

        folder_path = self.paths.dxf_root_dir / dxfrange.folder_name
        main_dwg = folder_path / f"{dxf_no}.dwg"

        return DXFFileRecord(
            dxf_no=dxf_no,
            folder_name=dxfrange.folder_name,
            folder_path=str(folder_path),
            main_dwg_path=str(main_dwg),
            has_main_dwg=main_dwg.is_file(),
        )

    def rebuild_files_index_full(
        self,
        progress_cb: Optional[Callable[[str], None]] = None,
    ) -> int:
        """
        Полный rebuild файлового DXF-индекса.
        """
        if not self.is_dxf_root_available():
            raise FileNotFoundError(f"DXF-Root nicht verfügbar: {self.paths.dxf_root_dir}")

        if not self._ranges:
            self.rebuild_ranges_from_workbook()

        result: dict[int, DXFFileRecord] = {}
        total = self.indexing.dxf.max_dxf_no - self.indexing.dxf.min_dxf_no + 1
        done = 0

        for dxf_no in range(self.indexing.dxf.min_dxf_no, self.indexing.dxf.max_dxf_no + 1):
            result[dxf_no] = self._build_file_record(dxf_no)
            done += 1

            if progress_cb and (done % 100 == 0 or done == total):
                progress_cb(f"DXF {dxf_no} geprüft ({done}/{total})")

        self._save_files_index_json(result)
        self._files_index = result
        return len(result)

    # --------------------------------------------------------
    # Search helpers
    # --------------------------------------------------------

    def _make_search_result(
        self,
        rec: DXFExcelRecord,
        file_rec: Optional[DXFFileRecord],
    ) -> DXFSearchResult:
        """
        Объединяет Excel-запись и файловую запись в DXFSearchResult.
        """
        return DXFSearchResult(
            dxf_no=rec.dxf_no,
            k_num=rec.k_num,
            wst=rec.wst,
            dicke_mm=rec.dicke_mm,
            ch_nr=rec.ch_nr,
            a_kn_brutto_qm=rec.a_kn_brutto_qm,
            laenge_zuschnitt_mm=rec.laenge_zuschnitt_mm,
            preis_pro_laenge_eur=rec.preis_pro_laenge_eur,
            main_dwg_path=file_rec.main_dwg_path if file_rec else "",
            has_main_dwg=file_rec.has_main_dwg if file_rec else False,
        )

    # --------------------------------------------------------
    # Public search API
    # --------------------------------------------------------

    def search_by_k_num(self, k_num: str) -> list[DXFSearchResult]:
        """
        Ищет все DXF-записи, связанные с указанным K-номером.
        """
        k_norm = self._normalize_k_num(k_num)
        if not k_norm:
            return []

        results: list[DXFSearchResult] = []
        for rec in self._by_k_num.get(k_norm, []):
            file_rec = self._files_index.get(rec.dxf_no)
            results.append(self._make_search_result(rec, file_rec))

        results.sort(key=lambda x: (x.dxf_no, x.k_num))
        return results

    def search_by_dxf_no(self, query: str) -> list[DXFSearchResult]:
        """
        Точный поиск по DXF-номеру.
        """
        raw = str(query).strip()
        if not raw or not raw.isdigit():
            return []

        dxf_no = int(raw)
        results: list[DXFSearchResult] = []

        for rec in self._by_dxf_no.get(dxf_no, []):
            file_rec = self._files_index.get(dxf_no)
            results.append(self._make_search_result(rec, file_rec))

        if not results:
            file_rec = self._files_index.get(dxf_no)
            if file_rec:
                results.append(DXFSearchResult(
                    dxf_no=dxf_no,
                    k_num="",
                    wst="",
                    dicke_mm=None,
                    ch_nr="",
                    a_kn_brutto_qm=None,
                    laenge_zuschnitt_mm=None,
                    preis_pro_laenge_eur=None,
                    main_dwg_path=file_rec.main_dwg_path,
                    has_main_dwg=file_rec.has_main_dwg,
                ))

        results.sort(key=lambda x: (x.dxf_no, x.k_num, x.main_dwg_path))
        return results

    def search_by_dxf_partial(self, query: str) -> list[DXFSearchResult]:
        """
        Частичный поиск по началу DXF-номера.
        """
        raw = str(query).strip()
        if not raw or not raw.isdigit():
            return []

        results: list[DXFSearchResult] = []
        matched_numbers = sorted(
            no for no in (set(self._by_dxf_no.keys()) | set(self._files_index.keys()))
            if str(no).startswith(raw)
        )

        for dxf_no in matched_numbers:
            excel_rows = self._by_dxf_no.get(dxf_no, [])
            file_rec = self._files_index.get(dxf_no)

            if excel_rows:
                for rec in excel_rows:
                    results.append(self._make_search_result(rec, file_rec))
            elif file_rec:
                results.append(DXFSearchResult(
                    dxf_no=dxf_no,
                    k_num="",
                    wst="",
                    dicke_mm=None,
                    ch_nr="",
                    a_kn_brutto_qm=None,
                    laenge_zuschnitt_mm=None,
                    preis_pro_laenge_eur=None,
                    main_dwg_path=file_rec.main_dwg_path,
                    has_main_dwg=file_rec.has_main_dwg,
                ))

        results.sort(key=lambda x: (x.dxf_no, x.k_num, x.main_dwg_path))
        return results

    def search(self, query: str) -> list[DXFSearchResult]:
        """
        Унифицированный поиск по DXF-запросу.

        Сейчас это partial search по началу номера.
        Используется из main_frame.py в режиме DXF.
        """
        return self.search_by_dxf_partial(query)


# ============================================================
# AppNr Repository
# ============================================================

class AppNrRepository:
    """
    Репозиторий серийных номеров аппаратов.

    Хранит:
    - _all_records : полный список записей
    - _by_prefix   : prefix -> list[AppNrRecord]
    - _by_k_code   : K-код -> list[str]
    """

    FULL_K_RE = re.compile(r"^K\d{5}$", re.IGNORECASE)

    def __init__(self, settings: AppSettings):
        self.settings = settings
        self.paths = settings.paths
        self.indexing = settings.indexing
        self.data_files = settings.data_files

        self._all_records: list[AppNrRecord] = []
        self._by_prefix: dict[str, list[AppNrRecord]] = {}
        self._by_k_code: dict[str, list[str]] = {}

        self.reload()

    # --------------------------------------------------------
    # Helpers
    # --------------------------------------------------------

    def _normalize_k_code(self, value) -> str:
        """
        Приводит значение к формату K12345.
        """
        if value is None:
            return ""

        s = str(value).strip().upper()
        if not s:
            return ""

        if re.fullmatch(r"\d{5}", s):
            return f"K{s}"

        if self.FULL_K_RE.fullmatch(s):
            return s

        return ""

    def _normalize_serial_text(self, value) -> str:
        """
        Нормализует серийный номер.
        """
        if value is None:
            return ""
        return str(value).strip()

    def _extract_prefix(self, serial_text: str) -> str:
        """
        Извлекает числовой префикс из серийного номера.

        Примеры:
        - 1234       -> 1234
        - 1234.1     -> 1234
        - 1234.1-2   -> 1234
        """
        s = serial_text.strip()
        if not s:
            return ""

        head = s.split(".", 1)[0].strip()
        m = re.match(r"^(\d+)", head)
        return m.group(1) if m else ""

    def is_excel_available(self) -> bool:
        """
        Проверяет доступность Excel-файла Apparate-Nr.
        """
        return self.paths.appnr_excel_file.exists() and self.paths.appnr_excel_file.is_file()

    # --------------------------------------------------------
    # JSON
    # --------------------------------------------------------

    def _load_raw(self) -> dict:
        path = self.data_files.appnr_index_json
        if not path.is_file():
            return {"items": []}

        try:
            with path.open("r", encoding="utf-8") as f:
                data = json.load(f)

            if not isinstance(data, dict) or "items" not in data:
                return {"items": []}

            return data
        except (OSError, json.JSONDecodeError) as e:
            logger.error(f"AppNrRepository._load_raw: {e}")
            return {"items": []}

    def _save_raw(self, data: dict) -> None:
        path = self.data_files.appnr_index_json
        path.parent.mkdir(parents=True, exist_ok=True)

        with path.open("w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)

    def reload(self) -> None:
        """
        Загружает индекс с диска в память.
        """
        data = self._load_raw()

        self._all_records = []
        self._by_prefix = {}
        self._by_k_code = {}

        for item in data.get("items", []):
            try:
                rec = AppNrRecord.from_dict(item)
            except (KeyError, TypeError, ValueError):
                continue

            self._all_records.append(rec)
            self._by_prefix.setdefault(rec.serial_prefix, []).append(rec)
            self._by_k_code.setdefault(rec.k_code, []).append(rec.serial_no)

        self._all_records.sort(key=lambda x: (x.serial_no, x.k_code))

        for prefix in self._by_prefix:
            self._by_prefix[prefix].sort(key=lambda x: (x.serial_no, x.k_code))

        for k_code in self._by_k_code:
            self._by_k_code[k_code] = sorted(set(self._by_k_code[k_code]))

    # --------------------------------------------------------
    # Rebuild
    # --------------------------------------------------------

    def rebuild_full(self) -> int:
        """
        Полный rebuild индекса App.Nr. из Excel.
        """
        if not self.is_excel_available():
            raise FileNotFoundError(f"Excel-Datei nicht gefunden: {self.paths.appnr_excel_file}")

        wb = load_workbook(
            self.paths.appnr_excel_file,
            data_only=True,
            read_only=True,
        )
        ws = wb.active

        items: list[AppNrRecord] = []
        first = True

        for row in ws.iter_rows(values_only=True):
            if first:
                first = False
                continue

            serial_no = self._normalize_serial_text(row[0] if len(row) > 0 else None)
            k_code = self._normalize_k_code(row[4] if len(row) > 4 else None)

            if not serial_no or not k_code:
                continue

            prefix = self._extract_prefix(serial_no)
            if not prefix:
                continue

            items.append(AppNrRecord(
                serial_no=serial_no,
                serial_prefix=prefix,
                k_code=k_code,
            ))

        wb.close()

        self._save_raw({"items": [x.to_dict() for x in items]})
        self.reload()
        return len(items)

    def rebuild_tail(self) -> int:
        """
        Хвостовой rebuild индекса App.Nr.

        Логика:
        - если индекса ещё нет -> полный rebuild
        - определяем максимальный известный serial_prefix
        - пересчитываем только записи с prefix >= max_prefix - tail_backtrack_prefix
        - старую "стабильную" часть индекса сохраняем без изменений
        """
        if not self._all_records:
            return self.rebuild_full()

        if not self.is_excel_available():
            raise FileNotFoundError(f"Excel-Datei nicht gefunden: {self.paths.appnr_excel_file}")

        tail_backtrack = self.indexing.appnr.tail_backtrack_prefix

        numeric_prefixes = [
            int(rec.serial_prefix)
            for rec in self._all_records
            if rec.serial_prefix.isdigit()
        ]

        if not numeric_prefixes:
            return self.rebuild_full()

        max_prefix = max(numeric_prefixes)
        limit = max(0, max_prefix - tail_backtrack)

        stable_part = [
            rec for rec in self._all_records
            if rec.serial_prefix.isdigit() and int(rec.serial_prefix) < limit
        ]

        wb = load_workbook(
            self.paths.appnr_excel_file,
            data_only=True,
            read_only=True,
        )
        ws = wb.active

        tail_part: list[AppNrRecord] = []
        first = True

        for row in ws.iter_rows(values_only=True):
            if first:
                first = False
                continue

            serial_no = self._normalize_serial_text(row[0] if len(row) > 0 else None)
            k_code = self._normalize_k_code(row[4] if len(row) > 4 else None)

            if not serial_no or not k_code:
                continue

            prefix = self._extract_prefix(serial_no)
            if not prefix or not prefix.isdigit():
                continue

            if int(prefix) >= limit:
                tail_part.append(AppNrRecord(
                    serial_no=serial_no,
                    serial_prefix=prefix,
                    k_code=k_code,
                ))

        wb.close()

        merged = stable_part + tail_part
        merged.sort(key=lambda x: (x.serial_no, x.k_code))

        self._save_raw({"items": [x.to_dict() for x in merged]})
        self.reload()
        return len(tail_part)

    # --------------------------------------------------------
    # Search
    # --------------------------------------------------------

    def search(self, query: str) -> list[AppNrRecord]:
        """
        Частичный поиск по началу полного серийного номера.
        """
        q = str(query).strip()
        if not q:
            return []

        q_upper = q.upper()
        results = [
            rec for rec in self._all_records
            if rec.serial_no.upper().startswith(q_upper)
        ]
        results.sort(key=lambda x: (x.serial_no, x.k_code))
        return results

    def get_serials_for_k(self, k_code: str) -> list[str]:
        """
        Возвращает все серийные номера для указанного K-кода.
        """
        return list(self._by_k_code.get(k_code.strip().upper(), []))
