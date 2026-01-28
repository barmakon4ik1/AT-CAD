#!Python3
# Программа вывода на экран готовности деталей лазерной резки в соответсвии с контрактом
# Version 1 / 05.07.2021

# импорт модуля openpyxl
import openpyxl, pprint
print('File DXF öffnen, bitte warten!')

# импорт модуля для красивых таблиц |"pip install prettytable"|
from prettytable import PrettyTable

#Укажите местоположение файла
path="G:\\Drawing\\DXF-LASER\\DXF-2017.xlsm"

# Чтобы открыть книгу создан объект рабочей книги
wb_obj = openpyxl.load_workbook(path)

# Получить активный лист
sheet = wb_obj.active

# Объекты ячейки имеют строку, столбец и координаты атрибутов,
# которые обеспечивают информацию о местоположении для ячейки.
# Примечание: первый ряд или Целое число столбца равно 1, а не 0.
# Объект ячейки создается с помощью метода cell () объекта листа

print('Zeilen lesen, bitte warten!')
while True:
    print()
    # print('Введите номер работы или любой нецифровой символ для выхода')
    print ('Bitte K-Nummer eingeben oder ein beliebiges nicht numerisches Zeichen um das Programm zu beenden')
    kom=input()
	
    table = PrettyTable()
	
    # Названия полей в заголовках
    table.field_names = ["DXF", "Wst", "Dicke", "Charge"]


    if kom.isalpha():
        break
    print('***********************************************************************')
    print("K" + str(kom) +":")
    print("DXF:   Material:  Dicke:  Charge-Nr.")

    for i in range(2, sheet.max_row + 1):
        if str(sheet.cell(row=i, column=2).value)==kom:
            # добавление строк данных
            table.add_row([sheet['A'+str(i)].value, sheet['D'+str(i)].value, sheet['E'+str(i)].value, sheet['F'+str(i)].value])
            # DXF=sheet['A'+str(i)].value   
            # Wst=sheet['D'+str(i)].value
            # Dicke=sheet['E'+str(i)].value
            # Charge=sheet['F'+str(i)].value
            # print(str(DXF) + "   " + str(Wst) + "      " + str(Dicke) + " мм   " + str(Charge))

    print(table)
	
    print('***********************************************************************')

    










